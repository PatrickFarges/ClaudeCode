r"""
build_index.py — Cross-reference Test Evidence + OneNote -> Excel maitre

Lit :
  - evidence_scan.csv (sortie de scan_evidence.py)
  - onenote_scan.csv  (sortie de scan_onenote.py)

Joint sur le ticket_id et produit sap_onenote_index.xlsx avec :
  - "Tickets" : ligne par ticket, hyperliens vers dossier Windows ET page OneNote
  - "OneNote orphelins" : pages OneNote avec ID mais sans dossier evidence
  - "Dossiers orphelins" : dossiers Tickets sans page OneNote
  - "OneNote sans ID" : pages OneNote sans CSxxx/CHGxxx (annotations utiles)
  - "Stats" : couverture, repartition par client/section

APP_VERSION 1.0.0 (2026-05-01) : version initiale
"""

import csv
from collections import defaultdict
from pathlib import Path
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

APP_VERSION = "1.0.0"

HERE = Path(__file__).parent
EVIDENCE_CSV = HERE / "evidence_scan.csv"
ONENOTE_CSV = HERE / "onenote_scan.csv"
OUT_XLSX = HERE / "sap_onenote_index.xlsx"

LINK_FONT = Font(color="0563C1", underline="single")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")


def read_csv(path: Path) -> list[dict]:
    if not path.exists():
        print(f"[ERREUR] {path} introuvable.")
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f, delimiter=";"))


def folder_to_uri(folder_path: str) -> str:
    """Transforme un chemin Windows en URI file: cliquable depuis Excel."""
    if not folder_path:
        return ""
    p = folder_path.replace("\\", "/")
    return "file:///" + quote(p, safe="/:")


def write_header(ws, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def autofit(ws, max_width: int = 60) -> None:
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def add_link_cell(ws, row: int, col: int, display: str, url: str) -> None:
    c = ws.cell(row=row, column=col, value=display)
    if url:
        c.hyperlink = url
        c.font = LINK_FONT


def build_tickets_sheet(wb: Workbook, evidence: list[dict],
                       onenote: list[dict]) -> tuple[set[str], set[str]]:
    """Onglet principal : un ticket par ligne, joint sur l'ID."""
    ws = wb.create_sheet("Tickets")

    # Index OneNote par ID (peut avoir plusieurs pages pour le meme ID)
    onenote_by_id: dict[str, list[dict]] = defaultdict(list)
    for r in onenote:
        if r["id"]:
            onenote_by_id[r["id"]].append(r)

    headers = [
        "ID", "Client", "Annee", "Folder name", "Folder (cliquable)",
        "Excel files (match ID)", "Page OneNote", "Section",
        "OneNote (cliquable)",
    ]
    write_header(ws, headers)

    matched_ids: set[str] = set()
    matched_pages: set[str] = set()
    row = 2
    for e in evidence:
        tid = e["id"]
        pages = onenote_by_id.get(tid, [])
        if pages:
            matched_ids.add(tid)
        # Si plusieurs pages OneNote pour le meme ID, on prend la 1ere et concatene
        page = pages[0] if pages else None
        if page:
            matched_pages.add(page["page_id"])

        ws.cell(row=row, column=1, value=tid)
        ws.cell(row=row, column=2, value=e["client"])
        ws.cell(row=row, column=3, value=e["annee"])
        ws.cell(row=row, column=4, value=e["folder_name"])
        add_link_cell(ws, row, 5, "[Ouvrir]", folder_to_uri(e["folder_path"]))
        ws.cell(row=row, column=6, value=e["excel_files"])
        if page:
            extra = f" (+{len(pages)-1} autres)" if len(pages) > 1 else ""
            ws.cell(row=row, column=7, value=page["page_title"] + extra)
            ws.cell(row=row, column=8, value=page["section"])
            add_link_cell(ws, row, 9, "[Ouvrir OneNote]", page["hyperlink"])
        else:
            ws.cell(row=row, column=7, value="(pas de page OneNote)")
        row += 1

    autofit(ws)
    return matched_ids, matched_pages


def build_onenote_orphans_sheet(wb: Workbook, onenote: list[dict],
                                matched_pages: set[str]) -> int:
    """Pages OneNote avec un ID mais aucun dossier evidence correspondant."""
    ws = wb.create_sheet("OneNote orphelins")
    headers = ["ID", "Page OneNote", "Section", "OneNote (cliquable)"]
    write_header(ws, headers)
    row = 2
    count = 0
    for r in onenote:
        if not r["id"]:
            continue
        if r["page_id"] in matched_pages:
            continue
        ws.cell(row=row, column=1, value=r["id"])
        ws.cell(row=row, column=2, value=r["page_title"])
        ws.cell(row=row, column=3, value=r["section"])
        add_link_cell(ws, row, 4, "[Ouvrir OneNote]", r["hyperlink"])
        row += 1
        count += 1
    autofit(ws)
    return count


def build_no_id_sheet(wb: Workbook, onenote: list[dict]) -> int:
    """Pages OneNote sans ID CS/CHG : annotations, procedures, navigation SAP."""
    ws = wb.create_sheet("OneNote sans ID")
    headers = ["Page OneNote", "Notebook", "Section", "OneNote (cliquable)"]
    write_header(ws, headers)
    row = 2
    count = 0
    for r in onenote:
        if r["id"]:
            continue
        ws.cell(row=row, column=1, value=r["page_title"])
        ws.cell(row=row, column=2, value=r["notebook"])
        ws.cell(row=row, column=3, value=r["section"])
        add_link_cell(ws, row, 4, "[Ouvrir OneNote]", r["hyperlink"])
        row += 1
        count += 1
    autofit(ws)
    return count


def build_stats_sheet(wb: Workbook, evidence: list[dict], onenote: list[dict],
                      matched_ids: set[str], onenote_orphans: int,
                      no_id: int) -> None:
    ws = wb.create_sheet("Stats", 0)  # premier onglet
    ws.title = "Stats"

    n_ev = len(evidence)
    n_on = len(onenote)
    n_on_with_id = sum(1 for r in onenote if r["id"])
    pct = lambda a, b: f"{(a / b * 100):.1f} %" if b else "-"

    rows = [
        ("Indicateur", "Valeur", "Commentaire"),
        ("Tickets evidence (dossiers)", n_ev, "depuis evidence_scan.csv"),
        ("Pages OneNote (apres filtre)", n_on, "depuis onenote_scan.csv"),
        ("Pages OneNote avec ID", n_on_with_id, ""),
        ("Pages OneNote sans ID", no_id, "annotations / proc / nav SAP"),
        ("", "", ""),
        ("Tickets evidence avec page OneNote", len(matched_ids),
         pct(len(matched_ids), n_ev) + " de couverture cote dossiers"),
        ("Tickets evidence sans page OneNote", n_ev - len(matched_ids), ""),
        ("OneNote orphelins (ID sans dossier)", onenote_orphans,
         pct(onenote_orphans, n_on_with_id) + " des pages avec ID"),
        ("", "", ""),
        ("Genere par", f"build_index.py v{APP_VERSION}", ""),
    ]
    for r_idx, row_data in enumerate(rows, 1):
        for c_idx, val in enumerate(row_data, 1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                c.font = HEADER_FONT
                c.fill = HEADER_FILL

    # Repartition par client (tickets evidence)
    by_client: dict[str, dict] = defaultdict(lambda: {"total": 0, "with_one": 0})
    for e in evidence:
        by_client[e["client"]]["total"] += 1
        if e["id"] in matched_ids:
            by_client[e["client"]]["with_one"] += 1

    start = len(rows) + 3
    ws.cell(row=start, column=1, value="Par client").font = Font(bold=True)
    headers2 = ("Client", "Tickets", "Avec OneNote", "Couverture")
    for c_idx, h in enumerate(headers2, 1):
        c = ws.cell(row=start + 1, column=c_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    r = start + 2
    for client, d in sorted(by_client.items(), key=lambda x: -x[1]["total"]):
        ws.cell(row=r, column=1, value=client)
        ws.cell(row=r, column=2, value=d["total"])
        ws.cell(row=r, column=3, value=d["with_one"])
        ws.cell(row=r, column=4, value=pct(d["with_one"], d["total"]))
        r += 1

    autofit(ws, max_width=80)


def main() -> None:
    print(f"build_index.py v{APP_VERSION}\n")
    evidence = read_csv(EVIDENCE_CSV)
    onenote = read_csv(ONENOTE_CSV)
    print(f"[INFO] {len(evidence)} dossiers evidence, {len(onenote)} pages OneNote")

    if not evidence or not onenote:
        print("[ERREUR] CSV manquants. Lancer scan_evidence.py + scan_onenote.py d'abord.")
        return

    wb = Workbook()
    wb.remove(wb.active)  # supprime la feuille par defaut

    matched_ids, matched_pages = build_tickets_sheet(wb, evidence, onenote)
    onenote_orphans = build_onenote_orphans_sheet(wb, onenote, matched_pages)
    no_id = build_no_id_sheet(wb, onenote)
    build_stats_sheet(wb, evidence, onenote, matched_ids, onenote_orphans, no_id)

    wb.save(OUT_XLSX)
    print(f"\n[OK] Index ecrit : {OUT_XLSX}")
    print(f"  Tickets matches OneNote : {len(matched_ids)} / {len(evidence)}")
    print(f"  OneNote orphelins       : {onenote_orphans}")
    print(f"  OneNote sans ID         : {no_id}")


if __name__ == "__main__":
    main()
