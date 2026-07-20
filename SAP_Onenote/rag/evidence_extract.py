#!/usr/bin/env python3
"""
evidence_extract.py — Extraction des dossiers "Test Evidence" (phase 2 du RAG SAP).

APP_VERSION = 0.1.0

Chaque *répertoire* qui contient directement un ou plusieurs classeurs Excel est
une "unité ticket". On ne garde que le **fichier Excel le plus récent** (mtime,
fiable : les fichiers ont conservé leur date d'origine à la copie) et on en extrait
le texte des onglets qui comptent, par ordre de valeur décroissante :

  1. "Change Logs"  → CE QUI A ÉTÉ MODIFIÉ (tables/transactions/rubriques SAP). Essentiel.
  2. "Analyse"      → l'analyse faite par l'AMO avant la modif.
  3. "Analyse HRO"  → l'analyse du problème par HRO (pourquoi le ticket a été ouvert).

Métadonnées : N° ticket (CS/CHG/INC/DFCT) depuis le nom du dossier/fichier, client
(dossier racine, code XXXFR, ou "Customer Name" du formulaire), tables SAP (T511K,
V_T5F…, Y00BA_…). Si aucun onglet cible n'existe (formulaires DFCT/INC purs), on
retombe sur les onglets formulaire pour que la recherche par N° ticket trouve quand même.

Lecture robuste : .xlsx/.xlsm via openpyxl, .xls via xlrd, avec bascule automatique
sur l'autre moteur en cas de mauvaise extension (fréquent : exports SAP déguisés).

Change-log
----------
0.1.0 (2026-07-20) : version initiale (phase 2).
"""
import os
import re
import warnings

warnings.simplefilter("ignore")  # openpyxl râle sur les styles/validations exotiques

import openpyxl  # noqa: E402
import xlrd      # noqa: E402

EXCEL_EXT = (".xlsx", ".xlsm", ".xls")
MAX_ROWS_PER_SHEET = 600      # au-delà, du bruit (dumps de données)
MAX_BODY_CHARS = 32000        # borne la taille d'un doc

# --- Identifiants tickets -----------------------------------------------------
ID_RE = re.compile(r"\b(CS\d{7}|CHG\d{7,10}|INC\d{7,}|DFCT\d{6,})\b", re.IGNORECASE)
# Tables / vues SAP. On EXCLUT les rubriques de paie (/T02, /LBD, M175…) qui ne sont
# pas des tables : une table commence par T5xx, T<3 chiffres>, V_… ou Y00BA_.
_TAB_PATTERNS = [
    re.compile(r"\b(?:V_)?T5[0-9A-Z][0-9A-Z_]{1,12}\b"),   # T511K, T512W, T5F1B, V_T5F1G1
    re.compile(r"\bT[0-9]{3}[0-9A-Z]{0,4}\b"),             # T511, T001P, T077S0
    re.compile(r"\bV_[0-9][0-9A-Z][0-9A-Z_]{2,12}\b"),     # V_512W_D, V_5F1C_C
    re.compile(r"\bY00BA[0-9A-Z_]*\b", re.IGNORECASE),
]

# --- Clients : canonique -> jetons de reconnaissance (majuscules) -------------
# On évite les codes de 3 lettres nus (COR, AST…) trop ambigus ; on garde les
# formes FR-suffixées et les noms complets. Le dossier racine sert d'ancre.
CLIENTS = {
    "RECKITT":     ["RECKITT", "RKTFR", "RCKTFR"],
    "LEO PHARMA":  ["LEO PHARMA", "LEOFR", "LEO FR"],
    "AKZO NOBEL":  ["AKZO NOBEL", "AKZO", "AKNFR"],
    "CORNING":     ["CORNING", "CORFR"],
    "ASTELLAS":    ["ASTELLAS", "ASTFR"],
    "ABBVIE":      ["ABBVIE", "ABVFR"],
    "ALCON":       ["ALCON"],
    "LONZA":       ["LONZA"],
    "BRIDGESTONE": ["BRIDGESTONE", "BSEFR"],
    "VUELING":     ["VUELING", "VUEFR"],
    "BUNGE":       ["BUNGE", "BNGFR"],
    "BNY":         ["BNY"],
    "EUROAPI":     ["EUROAPI"],
    "GFK":         ["GFK"],
    "HCC":         ["HCC"],
    "INO":         ["INO"],
    "TECHNIP":     ["TECHNIP"],
    "SOLVAY":      ["SOLVAY"],
}
_CLIENT_PATTERNS = [
    (name, re.compile(r"\b(?:" + "|".join(re.escape(t) for t in toks) + r")\b"))
    for name, toks in CLIENTS.items()
]
_CLIENT_FOLDERS = {name.upper(): name for name in CLIENTS}


def find_units(root):
    """root -> dict {répertoire: chemin de l'Excel le plus récent}."""
    units = {}
    for dp, _dn, fn in os.walk(root):
        xls = [f for f in fn
               if f.lower().endswith(EXCEL_EXT) and not f.startswith("~$")]
        if not xls:
            continue
        xls.sort(key=lambda f: os.path.getmtime(os.path.join(dp, f)), reverse=True)
        units[dp] = os.path.join(dp, xls[0])
    return units


def _classify(name):
    low = name.strip().lower()
    if "change log" in low:
        return "change_log"
    if "analy" in low and "hro" in low:
        return "hro"
    if "analy" in low:
        return "analyse"
    return None


def _is_context(name):      # formulaire de tête (Customer Name, Change Reason…)
    return "test evidence form" in name.strip().lower()


def _is_fallback_form(name):
    low = name.strip().lower()
    return (any(k in low for k in ("dfct form", "inc form", "incident form",
                                   "functional request"))
            or low in ("sheet1",))


def _want(name):
    return bool(_classify(name)) or _is_context(name) or _is_fallback_form(name)


def _row_text(row):
    cells = []
    for c in row:
        if c is None:
            continue
        s = str(c).strip()
        if s and s.lower() != "none":
            cells.append(s)
    return " | ".join(cells)


def read_sheets(path, want=_want, max_rows=MAX_ROWS_PER_SHEET):
    """Renvoie {nom_onglet: [lignes texte]} pour les onglets voulus.

    Bascule openpyxl<->xlrd selon l'extension, avec fallback sur l'autre moteur.
    """
    ext = os.path.splitext(path)[1].lower()
    readers = ["openpyxl", "xlrd"] if ext in (".xlsx", ".xlsm") else ["xlrd", "openpyxl"]
    last = None
    for r in readers:
        try:
            out = {}
            if r == "openpyxl":
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True,
                                            keep_links=False)
                try:
                    for name in wb.sheetnames:
                        if not want(name):
                            continue
                        ws = wb[name]
                        rows = []
                        for i, row in enumerate(ws.iter_rows(values_only=True)):
                            if i >= max_rows:
                                break
                            t = _row_text(row)
                            if t:
                                rows.append(t)
                        out[name] = rows
                finally:
                    wb.close()
            else:
                wb = xlrd.open_workbook(path, on_demand=True)
                for name in wb.sheet_names():
                    if not want(name):
                        continue
                    sh = wb.sheet_by_name(name)
                    rows = []
                    for ri in range(min(sh.nrows, max_rows)):
                        t = _row_text(sh.row_values(ri))
                        if t:
                            rows.append(t)
                    out[name] = rows
                wb.release_resources()
            return out
        except Exception as e:  # noqa: BLE001
            last = e
    raise last


def _kv_from_form(rows, key):
    """Cherche 'Key | Value | …' dans les lignes du formulaire, renvoie Value."""
    key = key.lower()
    for line in rows:
        parts = [p.strip() for p in line.split(" | ")]
        for i, p in enumerate(parts[:-1]):
            if p.lower().startswith(key):
                val = parts[i + 1].strip()
                if val and val.lower() not in (key, "value"):
                    return val
    return ""


def detect_client(relpath, filename, customer):
    """Détecte le client : dossier racine, sinon code/nom dans le chemin+customer."""
    top = relpath.split(os.sep)[0].upper().strip()
    if top in _CLIENT_FOLDERS:
        return _CLIENT_FOLDERS[top]
    # normalise les séparateurs (_ . -) en espaces pour que \b fonctionne
    # (ex. "LEO FR_Case" ou "RKTFR_FR11" ne cassent plus le motif)
    hay = re.sub(r"[^A-Z0-9 ]", " ", f"{relpath} {filename} {customer}".upper())
    for name, pat in _CLIENT_PATTERNS:
        if pat.search(hay):
            return name
    return ""


def _extract_ids(*texts):
    ids = []
    seen = set()
    for t in texts:
        for m in ID_RE.finditer(t or ""):
            v = m.group(0).upper()
            if v not in seen:
                seen.add(v)
                ids.append(v)
    return ids


def _extract_tables(text):
    out = set()
    for pat in _TAB_PATTERNS:
        for m in pat.finditer(text):
            out.add(m.group(0).upper())
    return sorted(out)


def extract_unit(dirpath, xls_path, root):
    """Extrait un doc structuré à partir d'un répertoire-ticket.

    Renvoie un dict prêt à insérer, ou None si le fichier est illisible.
    """
    relpath = os.path.relpath(dirpath, root)
    dirname = os.path.basename(dirpath.rstrip(os.sep))
    filename = os.path.basename(xls_path)

    try:
        sheets = read_sheets(xls_path)
    except Exception as e:  # noqa: BLE001
        return {"error": repr(e)[:120]}

    cl, an, hro, fb, tef = [], [], [], [], []
    for name, rows in sheets.items():
        if not rows:
            continue
        k = _classify(name)
        block = "\n".join(rows)
        if k == "change_log":
            cl.append(block)
        elif k == "analyse":
            an.append(block)
        elif k == "hro":
            hro.append(block)
        elif _is_context(name):
            tef.append(rows)          # gardé pour Customer/Change Reason
        elif _is_fallback_form(name):
            fb.append(block)

    customer = change_reason = ""
    for rows in tef:
        customer = customer or _kv_from_form(rows, "customer name")
        change_reason = change_reason or _kv_from_form(rows, "change reason")

    # Titre = nom du dossier (déjà le "<ID> - <description>" dans la majorité des cas)
    title = dirname if dirname and not re.fullmatch(r"\d{2}-\d{2}-\d{4}", dirname) \
        else os.path.splitext(filename)[0]

    parts = []
    if change_reason:
        parts.append(f"CONTEXTE (change reason) : {change_reason}")
    if cl:
        parts.append("== CHANGE LOG ==\n" + "\n".join(cl))
    if an:
        parts.append("== ANALYSE (AMO) ==\n" + "\n".join(an))
    if hro:
        parts.append("== ANALYSE HRO ==\n" + "\n".join(hro))
    if not (cl or an or hro) and fb:
        parts.append("== FORMULAIRE ==\n" + "\n".join(fb))
    body = "\n\n".join(parts)[:MAX_BODY_CHARS]

    ids = _extract_ids(dirname, filename, body)
    client = detect_client(relpath, filename, customer)
    tables = _extract_tables(f"{title}\n{body}")

    return {
        "title": title,
        "ticket_ids": " ".join(ids),
        "client": client,
        "tables_sap": " ".join(tables),
        "evidence_path": relpath,
        "excel_file": filename,
        "has_change_log": bool(cl),
        "has_analyse": bool(an),
        "has_hro": bool(hro),
        "body": body,
    }


if __name__ == "__main__":
    import sys
    import json
    root = os.environ.get("EVIDENCE_ROOT",
                          "/media/red/Samsung2TB/SAP_KB/TestEvidence")
    if len(sys.argv) > 1:
        # debug : extrait un répertoire précis
        d = sys.argv[1]
        units = find_units(d) if os.path.isdir(d) else {}
        if not units and os.path.isdir(d):
            print("aucun Excel dans", d)
        for dp, p in list(units.items())[:5]:
            print("=" * 80)
            r = extract_unit(dp, p, root)
            r["body"] = r.get("body", "")[:1500]
            print(json.dumps(r, ensure_ascii=False, indent=2))
    else:
        units = find_units(root)
        print(f"{len(units)} répertoires-tickets sous {root}")
