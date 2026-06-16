"""
generate_wtc_absences.py — Générateur du Wagetype Catalog Absences (multi-clients)

Modèle : `WTCA reference.xlsx` — template Strada déjà finalisé (logo + couleurs
Strada en place, bloc des 11 classes d'absences P→Z déployé, onglets dans l'ordre
voulu : Absences-Présences / Parameter (masqué) / Wagetype Catalog Absence_Data /
Calendrier / Aide remplissage). Le code ne fait plus AUCUN rebranding visuel : il
se contente de vider les données d'exemple et d'injecter celles du client.
Cibles : un fichier .xlsx par client (AKN, ABV, …) configuré dans CLIENT_CONFIGS,
ou un fichier choisi par l'HRO via l'interface dossier (build_dir_config).

Le template `WTCA reference.xlsx` a lui-même été généré par une version antérieure
de ce script (branding NGA→Strada appliqué une fois pour toutes) puis nettoyé à la
main par Pat (onglets "Taux de valorisation" et "Sheet1" supprimés). Les anciens
fichiers `GFK Original Absence Catalog.xls/.xlsx` ne sont plus utilisés ; les
fonctions de rebranding (replace_logos, patch_colors, convert_xls_to_xlsx,
load_template) sont conservées en réserve mais ne sont plus appelées par le flux
de génération — elles ne resservent qu'à régénérer un template depuis zéro.

Phases :
  Phase 2.1 — Injection des vraies données client (T554S + T554T, langue F)
  Phase 2.2 — Remplissage cols D-I via T554C → T512T + T511 (unités)
  Phase 2.3 — Ré-application des listes déroulantes (data validations) aux
              vraies lignes de données

Changelog :
  v0.1.0 — Phase 1 : copie + rebranding visuel Strada
  v0.2.0 — Phase 2.0 : onglet <CLIENT>_Data avec les catégories d'absences
  v0.3.0 — Phase 2.1 : injection des vraies données dans l'onglet principal
  v0.4.0 — Phase 2.2 : remplissage cols D, E, F, G, H, I via T554C/T512T/T511
  v0.5.0 — Multi-clients : refactoring pour supporter ABV en plus de AKN.
           CLI `--client AKN|ABV`. Gestion de la colonne 'L.HCM' (AKN)
           OU 'GrPay' (ABV) selon la version de dump SAP. Tous les paramètres
           (GrSdP, GrPay, dossier, fichiers, sortie) sont dans CLIENT_CONFIGS.
  v0.6.0 — Remplissage cols J (Nombre) et K (Montant) via T511 col J ('C').
           Logique : rubrique recherchée dans T511.C = WTC col A → D → G ;
           la valeur trouvée en T511 col J (caractère ' ', 'A', '0'..'9') est
           traduite via NOMBRE_MONTANT_MAP en texte 'soit nbr / aucun nbr /
           oblig. / facult. / ou nbr' et symétrique côté montant.
  v0.7.0 — Inversion ordre de lookup : D → G → A (les CatAbsP en col A peuvent
           collisionner par hasard avec des rubriques de paie dans T511, donc on
           privilégie les vraies rubriques Paiement/Retenue). Si T511 col C est
           vide sur la ligne trouvée → cols J et K du WTC laissées vides (cas
           fréquent sur rubriques 9000+ chez certains clients).
  v0.8.0 — Remplissage col L (Unité de temps — sous-section "Pilotage Paie") :
           même lookup que F et I (via T511.UnT) mais en partant de la rubrique
           Paiement (col D), avec fallback sur la rubrique Retenue (col G).
           La col A (CatAbsP) n'est PAS utilisée ici puisque L est dans la zone
           Pilotage Paie — basée uniquement sur les rubriques de paie.
  v0.9.0 — Cols P/Q/R/S (pénalisants) via les 10 colonnes CLABS de T554C.
           Pour chaque règle de valorisation, on lit la liste des CLABS et on
           coche ■ dans la col WTC correspondante si la classe figure dedans :
             P = CLABS 30 (Pénalisant 13ème mois)
             Q = CLABS 40 (Pénalisant Prime de Vacances)
             R = CLABS 50 (Pénalisant Prime d'ancienneté)
             S = CLABS 70 (Pénalisant RATP → transport)
           T554E sert seulement de référentiel CLABS → libellé (non utilisée
           pour la jointure, qui se fait via Règle de valorisat. dans T554C).
  v1.0.0 — TOUTES les classes d'absences de T554E (demande client ABVIE).
           Le bloc pénalisant passe de 4 colonnes (P-S) à 1 colonne par classe
           T554E, dans l'ordre des CLABS (1,10,20,30,40,50,60,70,80,90,91) :
             P=1 (Congé)        Q=10 (Maternité/Pat.)  R=20 (Maladie)
             S=30 (Pén.13ème)   T=40 (Pén.Pr.Vac.)     U=50 (Pén.Pr.Anc.)
             V=60 (Base Stag.)  W=70 (abs n.payée RATP) X=80 (abs payée RATP)
             Y=90 (Base Appr.)  Z=91 (Abs n.payées DSN)
           Les colonnes historiques à droite des pénalisants (Additional
           Comments, WD absence name, Customizing…) ne sont PAS écrasées mais
           DÉCALÉES de +N colonnes (N = nb classes − 4) vers la droite, avec
           leurs valeurs/styles/largeurs/fusions. Les entêtes ligne 7 reprennent
           le libellé 'Classe' de T554E. T554E est désormais chargée et pilote
           la structure du bloc (clé : GrPay). Clients dont T554E n'a pas été
           dumpée (ex. AKN) → repli sur DEFAULT_PEN_CLASSES (les 11 classes,
           identiques pour tous les clients GrPay=06).
  v1.0.1 — Fix : dé-masquage des lignes du template. Le template GFK hérite de
           ~93 lignes masquées (plan Excel du catalogue GFK d'origine, lignes
           12-156). Les lignes d'absence injectées tombant dessus étaient
           invisibles à l'écran (ex. CatAbsP 1202-1206 cachées entre 1201 et
           1207) bien que présentes dans le fichier.
  v1.0.2 — On ne reconduit AUCUN masquage du template : toutes les lignes de
           l'onglet principal sont rendues visibles (hidden=False sur tout
           ws.row_dimensions, pas seulement les lignes remplies). Le masquage
           HRO d'origine cachait des codes d'absence valides ; si le client veut
           masquer des codes inutilisés, ce sera à lui de le faire.
  v1.1.0 — Override métier ABVIE "Temps partiel thérapeutique" (CatAbsP
           1200-1209), piloté par cfg['therapeutic_partial'] :
             - col D (Rubrique Paiement) = texte fixe 'IT 2010' (rubriques de
               paie dans l'infotype 2010, envoyées par le client ; HRO ne veut
               pas de n° qui laisserait croire à un déclenchement auto) ;
             - col G (Rubrique Retenue) recopiée du code source 1200 (3170) +
               libellé/unité, pour 1201-1209 quand la retenue est vide.
           Sans config (ex. AKN) → aucun effet.
  v1.2.0 — Refactor pour l'interface HRO (wtc_absences_gui.py, Tkinter) :
             - generate_from_config(cfg) : cœur réutilisable, renvoie des stats ;
               generate(client) n'est plus qu'un wrapper CLI ;
             - build_dir_config(input_dir, output_path) + scan_tables_dir() :
               config dynamique depuis un dossier de tables choisi par l'HRO
               (params France par défaut, détection auto des fichiers de tables) ;
             - add_data_sheet utilise cfg['label'] (nom d'onglet assaini) ;
             - generate_from_config crée le dossier de sortie au besoin.
           La case 'fichier WTC antérieur à réviser' (cfg['revise_previous']) est
           prévue mais sans effet (reprise des colonnes manuelles HRO à venir).
  v1.3.0 — Bascule du template GFK vers `WTCA reference.xlsx` (template Strada
           déjà finalisé) :
             - plus d'appel replace_logos / patch_colors (template déjà brandé ;
               évite aussi d'écraser par erreur les captures de "Aide remplissage") ;
             - bloc des classes d'absences déjà à 11 colonnes dans le template →
               PEN_TEMPLATE_COUNT 4→11, TRAILING_END_COL Y(25)→AF(32),
               DATA_ROW_END 156→232 (vide toutes les données d'exemple) ;
             - l'onglet de données à plat est désormais peuplé EN PLACE sous son
               nom fixe 'Wagetype Catalog Absence_Data' (position conservée :
               3e onglet) au lieu de créer un onglet '<label>_Data' en fin de
               classeur ;
             - ré-application des listes déroulantes (data validations) du template
               aux vraies lignes de données (col C=Used ■/□, F/I=unité de temps,
               L=Heure/Jour, P→Z=■/□). Les listes TH (Taux horaire, col N) et TJ
               (Taux jour, col O) sont IGNORÉES car leurs named ranges pointent sur
               #REF! dans le template (l'onglet 'Taux de valorisation' qui les
               définissait a été supprimé) — à recréer côté template si besoin.
           Les onglets Parameter (masqué) et Aide remplissage sont recopiés tels
           quels (jamais modifiés). L'onglet Calendrier était lui aussi recopié
           tel quel jusqu'en v1.3.0 ; depuis v1.4.0 il est généré (voir ci-dessous).
  v1.4.0 — Génération automatique de l'onglet 'Calendrier' depuis T554S.
           Le Calendrier liste les TYPES d'absence : champ T554S.TypAb (code à
           2 lettres : AI, MA, CP, RT…) + libellé (T554S 'Texte cat. prés./abs.').
           TypAb n'est renseigné que sur certaines lignes "représentatives" de
           T554S → on ne garde que les lignes où TypAb ≠ vide. Filtre GrSdP =
           cfg['grsdp'] (libellés alors dans la langue du grouping, FR pour
           GrSdP=6). Dédup sur la 1ʳᵉ occurrence, tri alphabétique par code.
           Le bloc vert clair 'Entrées spécifiques' (saisie manuelle HRO) est
           conservé et réinséré juste sous les données.
           ⚠ T508A (règles de plan de roulement) n'a RIEN à voir avec le
           Calendrier : c'était une fausse piste (son col 'ID cal. jours fériés'
           contient des codes pays type AT/CH qui ressemblent par hasard aux
           codes TypAb).
"""
from __future__ import annotations

APP_VERSION = "1.4.0"

import argparse
import shutil
import subprocess
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# ---- Chemins communs ------------------------------------------------------

ROOT = Path(__file__).resolve().parent
# Template de référence actuel : déjà brandé Strada (logo + couleurs), bloc des
# 11 classes d'absences déployé (P→Z), onglets dans l'ordre voulu. C'est le point
# de départ de toute génération depuis la v1.3.0.
WTCA_TEMPLATE = ROOT / "WTCA reference.xlsx"
# Anciens modèles GFK — conservés pour mémoire / régénération éventuelle d'un
# template depuis zéro, mais PLUS utilisés par le flux de génération courant.
GFK_TEMPLATE = ROOT / "GFK Original Absence Catalog.xlsx"
GFK_XLS_LEGACY = ROOT / "GFK Original Absence Catalog.xls"
LOGO_STRADA = ROOT / "logo_strada.png"

# Mapping numéros → libellés d'unités de temps (T511 col "UnT" → humain)
UNITS_MAP_FILE = ROOT / "numeros_vs_unités"

# ---- Configurations clients -----------------------------------------------

# Chaque client a son dossier de dumps SAP. Les noms de fichiers peuvent varier
# en casse (T554S.xlsx vs t554s.XLSX) selon la version du dump.
CLIENT_CONFIGS: dict[str, dict] = {
    "AKN": {
        "dir": ROOT / "AKN_17.05.2026",
        "tables": {  # case-insensitive lookup via _resolve_table()
            "T554S": "T554S.xlsx", "T554T": "T554T.xlsx", "T554C": "T554C.xlsx",
            "T511": "T511.xlsx", "T512T": "T512T.xlsx",
            "Y00BA": "Y00BA_TAB_COMPAN.xlsx",
        },
        "grsdp": "6",       # Groupe subdivisions personnel
        "lang_fr": "F",     # Code langue français
        "grpay": "06",      # Groupe de paie / L.HCM
        "mdt": "984",       # Mandant SAP
        "output": ROOT / "AKN France Absences-Presences Catalogue.xlsx",
        "title_replace": ("FR Absences", "AKN Absences"),
        "label": "AKN",     # nom de l'onglet Data + titre
    },
    "ABV": {
        "dir": ROOT / "ABV",
        "tables": {
            "T554S": "t554s.XLSX", "T554T": "t554t.XLSX", "T554C": "t554c.XLSX",
            "T554E": "T554E.XLSX",
            "T511": "t511.XLSX", "T512T": "T512T.XLSX",
            "Y00BA": "Y00BA_TAB_COMPAN.XLSX",
        },
        "grsdp": "6",       # France (LCC=FR1)
        "lang_fr": "F",
        "grpay": "06",      # ABV mappé à GrPay=06 dans Y00BA
        "mdt": None,        # ABV : pas de filtre Mdt (dumps sans colonne fiable)
        "output": ROOT / "ABV WTC.xlsx",
        "title_replace": ("FR Absences", "ABV Absences"),
        "label": "ABV",     # nom de l'onglet Data + titre
        # Override métier ABVIE — "Temps partiel thérapeutique" (CatAbsP 1200-1209) :
        #   col D = 'IT 2010' (rubriques de paie dans l'infotype 2010, envoyées
        #           par le client → HRO ne veut PAS de n° de rubrique) ;
        #   col G = recopie de la rubrique de retenue du code 1200 (3170) pour
        #           1201-1209 si vide (retenue déclenchée automatiquement).
        "therapeutic_partial": {
            "codes": [str(c) for c in range(1200, 1210)],  # 1200..1209
            "retenue_source": "1200",
            "payment_label": "IT 2010",
        },
    },
}

# Palette Strada (déjà utilisée dans patch_colors)
STRADA_DARK = "FF084028"   # vert très foncé (logo + bandeau titre)
STRADA_MED = "FF207F4F"    # vert moyen (entêtes catégorie)
STRADA_LIGHT = "FF18D878"  # vert clair
YELLOW = "FFFFFF99"        # jaune données (palette GFK conservée)
GREY_LIGHT = "FFC0C0C0"    # gris clair entêtes secondaires
GREY_MED = "FF969696"      # gris moyen entêtes groupes
WHITE = "FFFFFFFF"
BLACK = "FF000000"

# Mise en forme onglet principal
MAIN_SHEET = "Absences-Présences"
DATA_SHEET = "Wagetype Catalog Absence_Data"  # onglet de données à plat (nom fixe)
CALENDAR_SHEET = "Calendrier"      # onglet des types d'absence (T554S.TypAb)
CAL_EXTRA_ROWS = 7                 # lignes vides "Entrées spécifiques" sous les données
CAL_GREEN = "FFCCFFCC"             # vert clair du bloc de saisie manuelle (réf. template)
HEADER_ROW_END = 7        # lignes 1-7 = entêtes (à conserver)
DATA_ROW_START = 8        # première ligne de données
DATA_ROW_END = 232        # dernière ligne d'exemple à vider dans WTCA reference
MAIN_COL_COUNT = 32       # colonnes utiles A-AF dans le template WTCA reference

# Géométrie du bloc "pénalisants" / classes d'absences (T554E).
# Le template WTCA reference a DÉJÀ 11 colonnes de classes (P→Z) ; pour un client
# France (11 classes), shift = 0 → restructure_penalisant_block ne fait que
# réécrire les libellés d'entête (no-op si identiques). Un client à N≠11 classes
# décalerait les colonnes "trailing" comme avant.
PEN_START_COL = 16        # P — 1re colonne de classe d'absence
PEN_TEMPLATE_COUNT = 11   # le template WTCA reference a 11 colonnes (P→Z)
TRAILING_END_COL = 32     # AF — dernière colonne "trailing" à décaler du template
                          # (AA..AF : Additional Comments, WD absence name, Customizing…)

# Repli pour les clients dont T554E n'a pas été dumpée (ex. AKN). Les classes
# d'absences sont les MÊMES pour tous les clients GrPay=06 (confirmé par Pat),
# donc on reproduit ici les 11 classes T554E. Si un client fournit son T554E,
# load_t554e() prime (libellés exacts du dump). Ordre = CLABS croissants.
# Format identique à load_t554e() : liste ordonnée de {'clabs', 'label'}.
DEFAULT_PEN_CLASSES = [
    {'clabs': '1',  'label': 'Congé'},
    {'clabs': '10', 'label': 'Maternité/Paternité'},
    {'clabs': '20', 'label': 'Maladie'},
    {'clabs': '30', 'label': 'Pénalisant 13ème mois'},
    {'clabs': '40', 'label': 'Pénalisant Prime Vacances'},
    {'clabs': '50', 'label': 'Pénalisant Prime Ancienneté'},
    {'clabs': '60', 'label': 'Pénalisant Base Stagiaire'},
    {'clabs': '70', 'label': 'abs non payée Impact  RATP'},
    {'clabs': '80', 'label': 'abs  payée Impact  RATP'},
    {'clabs': '90', 'label': 'Pénalisant Base Apprenti'},
    {'clabs': '91', 'label': 'Absences non payées DSN'},
]

# Fichier intermédiaire (conversion .xls → .xlsx via libreoffice)
TMP_DIR = ROOT / ".tmp_build"
GFK_XLSX = TMP_DIR / "GFK Original Absence Catalog.xlsx"

# ---- Mapping T511.C → texte (Nombre, Montant) -----------------------------
# Le caractère trouvé en T511 col J (entête 'C') sur la ligne dont la rubrique
# correspond donne le couple (texte_Nombre, texte_Montant) à écrire dans les
# cols J et K du WTC. Référence : fichier 'nombre_montant' (racine projet).
# Note : un T511.C vide → J et K du WTC laissés vides (géré en amont).
NOMBRE_MONTANT_MAP: dict[str, tuple[str, str]] = {
    'A': ('soit nbr',  'soit mnt'),
    '0': ('aucun nbr', 'aucun mnt'),
    '1': ('aucun nbr', 'oblig.'),
    '2': ('oblig.',    'aucun mnt'),
    '3': ('ou nbr',    'au moins le mnt'),
    '4': ('facult.',   'oblig.'),
    '5': ('oblig.',    'facult.'),
    '6': ('oblig.',    'oblig.'),
    '7': ('facult.',   'facult.'),
    '8': ('aucun nbr', 'aucun mnt'),
    '9': ('aucun nbr', 'facult.'),
}


# ---- Palette de substitution NGA → STRADA ----------------------------------
# Format hex AARRGGBB (alpha = FF). openpyxl utilise des hex en argb.
COLOR_MAP = {
    "FF660066": "FF084028",  # mauve très foncé → vert foncé Strada
    "FF993366": "FF207F4F",  # mauve bordeaux  → vert moyen Strada
    "FF800080": "FF18D878",  # violet          → vert clair Strada
}


def convert_xls_to_xlsx(xls_path: Path, out_dir: Path) -> Path:
    """Convertit le .xls en .xlsx via LibreOffice headless (Linux uniquement).
    Renvoie le chemin du .xlsx produit. Utilisé pour générer le template
    une fois pour toutes ; sur Windows on utilise le .xlsx pré-converti."""
    out_dir.mkdir(parents=True, exist_ok=True)
    print(f"      Conversion .xls → .xlsx via libreoffice…")
    result = subprocess.run(
        ["libreoffice", "--headless", "--convert-to", "xlsx",
         str(xls_path), "--outdir", str(out_dir)],
        capture_output=True, text=True, timeout=120,
    )
    if result.returncode != 0:
        raise RuntimeError(f"libreoffice conversion failed: {result.stderr}")
    produced = out_dir / (xls_path.stem + ".xlsx")
    if not produced.exists():
        raise FileNotFoundError(f"Conversion result not found: {produced}")
    return produced


def load_template(template_path: Path, xls_fallback: Path):
    """Charge le template .xlsx. Si absent, tente de le générer depuis le .xls
    via LibreOffice (uniquement disponible sur Linux/Mac avec LibreOffice installé)."""
    if template_path.exists():
        print(f"[1/5] Chargement template {template_path.name}…")
        return openpyxl.load_workbook(template_path)
    if not xls_fallback.exists():
        raise FileNotFoundError(
            f"Ni le template .xlsx ({template_path}) ni le .xls de fallback "
            f"({xls_fallback}) ne sont disponibles."
        )
    print(f"[1/5] Template .xlsx absent — conversion depuis .xls…")
    TMP_DIR.mkdir(parents=True, exist_ok=True)
    converted = convert_xls_to_xlsx(xls_fallback, TMP_DIR)
    shutil.copy(converted, template_path)
    return openpyxl.load_workbook(template_path)


def patch_colors(wb) -> tuple[int, int]:
    """Substitue les couleurs mauves NGA par les couleurs vertes Strada
    sur toutes les feuilles. Renvoie (nb_cellules_fill_patches, nb_cellules_font_patches)."""
    print(f"[3/5] Substitution des couleurs NGA → Strada…")
    fill_count = 0
    font_count = 0
    for sname in wb.sheetnames:
        ws = wb[sname]
        for row in ws.iter_rows():
            for cell in row:
                # Fill background
                try:
                    if cell.fill and cell.fill.patternType and cell.fill.fgColor:
                        rgb = cell.fill.fgColor.rgb
                        if rgb and str(rgb) in COLOR_MAP:
                            new_rgb = COLOR_MAP[str(rgb)]
                            cell.fill = PatternFill(
                                patternType=cell.fill.patternType,
                                fgColor=new_rgb,
                                bgColor=cell.fill.bgColor.rgb if cell.fill.bgColor else "00000000",
                            )
                            fill_count += 1
                except (AttributeError, TypeError):
                    pass
                # Font color
                try:
                    if cell.font and cell.font.color and cell.font.color.rgb:
                        rgb = cell.font.color.rgb
                        if rgb and str(rgb) in COLOR_MAP:
                            new_rgb = COLOR_MAP[str(rgb)]
                            new_font = Font(
                                name=cell.font.name,
                                size=cell.font.size,
                                bold=cell.font.bold,
                                italic=cell.font.italic,
                                underline=cell.font.underline,
                                color=new_rgb,
                            )
                            cell.font = new_font
                            font_count += 1
                except (AttributeError, TypeError):
                    pass
    print(f"      → {fill_count} fills + {font_count} fontes patchées")
    return fill_count, font_count


def replace_logos(wb, logo_path: Path) -> int:
    """Remplace toutes les images des onglets par le logo Strada.
    Les anciens logos NGA sont supprimés et remplacés à la même position/taille
    (avec ratio préservé)."""
    print(f"[2/5] Remplacement des logos NGA → Strada…")
    if not logo_path.exists():
        raise FileNotFoundError(f"Logo Strada introuvable : {logo_path}")

    replaced = 0
    for sname in wb.sheetnames:
        ws = wb[sname]
        old_images = list(ws._images)
        if not old_images:
            continue
        print(f"      Sheet {sname!r}: {len(old_images)} image(s) à remplacer")
        # Capture les anchors avant de vider la liste
        anchors = [img.anchor for img in old_images]
        # Conserver les dimensions cible de chaque image originale (en EMU)
        # (openpyxl Image utilise width/height en pixels via PIL)
        from PIL import Image as PILImage
        ref_dims = []
        for img in old_images:
            # essayer de récupérer width/height ; fallback sur défaut
            w = getattr(img, "width", None)
            h = getattr(img, "height", None)
            ref_dims.append((w, h))
        # Vider la liste interne
        ws._images = []
        # Remettre une image Strada par anchor
        for anchor, (w, h) in zip(anchors, ref_dims):
            new_img = XLImage(str(logo_path))
            # Respect du ratio original du logo Strada (496×103)
            if w and h:
                # forcer à la même surface que l'original mais avec le ratio Strada
                ratio = 496 / 103
                # si on connait h, on adapte w pour conserver le ratio
                new_img.height = h
                new_img.width = int(h * ratio)
            new_img.anchor = anchor
            ws.add_image(new_img)
            replaced += 1
    print(f"      → {replaced} image(s) substituée(s)")
    return replaced


def load_units_map() -> dict[str, str]:
    """Lit numeros_vs_unités → dict { '001': 'Heures', '010': 'Jours', ... }.
    Ignore les 2 premières lignes d'en-tête explicatives (préfixées par '*')."""
    print(f"[4/5 c] Chargement mapping unités de temps ({UNITS_MAP_FILE.name})…")
    out: dict[str, str] = {}
    for raw in UNITS_MAP_FILE.read_text(encoding='utf-8').splitlines():
        line = raw.strip()
        if not line or line.startswith('*') or '=' not in line:
            continue
        # format possible "<n>\t<code>=<libellé>" ou "<code>=<libellé>"
        parts = line.split('\t', 1)
        kv = parts[-1].strip()
        code, _, label = kv.partition('=')
        code = code.strip()
        label = label.strip()
        if code and label:
            out[code] = label
    print(f"      → {len(out)} unités mappées")
    return out


def _resolve_table(cfg: dict, table_name: str) -> Path:
    """Renvoie le chemin de la table en gérant la casse (T554S.xlsx vs t554s.XLSX)."""
    p = cfg["dir"] / cfg["tables"][table_name]
    if p.exists():
        return p
    target = cfg["tables"][table_name].lower()
    for f in cfg["dir"].iterdir():
        if f.name.lower() == target:
            return f
    raise FileNotFoundError(f"Table {table_name} introuvable dans {cfg['dir']}")


def _idx(hdr, *names) -> int | None:
    """Renvoie l'index de la première colonne trouvée parmi 'names', ou None."""
    for n in names:
        if n in hdr:
            return hdr.index(n)
    return None


def _row_matches(r, idx_mdt, mdt_val, idx_grpay, grpay_val) -> bool:
    """Helper de filtrage Mdt + L.HCM/GrPay (chacun optionnel)."""
    if idx_mdt is not None and mdt_val is not None and r[idx_mdt] != mdt_val:
        return False
    if idx_grpay is not None and r[idx_grpay] != grpay_val:
        return False
    return True


def load_t511(cfg: dict) -> dict[str, dict]:
    """Charge T511 → dict { rubrique: {'unt': ..., 'c': ...} }.

    Filtre Mdt (si présent et défini) + L.HCM ou GrPay (selon la version du
    dump). Garde la version la plus récente (date Fin la plus haute).

    Le champ 'c' (entête T511 col J = 'C') sert au remplissage des cols J et K
    du WTC via NOMBRE_MONTANT_MAP. Espace et chaîne vide y sont normalisés à ' '.
    """
    table = _resolve_table(cfg, "T511")
    print(f"[4/5 d] Chargement T511 ({table.name})…")
    wb = openpyxl.load_workbook(table, data_only=True, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    hdr = next(rows)
    iM = _idx(hdr, 'Mdt')
    iG = _idx(hdr, 'L.HCM', 'GrPay')
    iR = hdr.index('Rubrique')
    iU = hdr.index('UnT')
    iC = hdr.index('C')
    iF = hdr.index('Fin')
    by_rub: dict[str, tuple] = {}
    for r in rows:
        if not _row_matches(r, iM, cfg["mdt"], iG, cfg["grpay"]):
            continue
        rub = r[iR]
        if not rub:
            continue
        fin = r[iF]
        unt = r[iU] or ''
        c_raw = r[iC]
        c = c_raw if isinstance(c_raw, str) and c_raw != '' else ' '
        prev = by_rub.get(rub)
        if prev is None or (fin and (not prev[2] or fin > prev[2])):
            by_rub[rub] = (unt, c, fin)
    wb.close()
    out = {k: {'unt': v[0], 'c': v[1]} for k, v in by_rub.items()}
    print(f"      → {len(out)} rubriques T511 trouvées")
    return out


def load_t512t(cfg: dict) -> dict[str, str]:
    """Charge T512T → dict { rubrique: libellé FR }."""
    table = _resolve_table(cfg, "T512T")
    print(f"[4/5 e] Chargement T512T ({table.name}, Langue={cfg['lang_fr']})…")
    wb = openpyxl.load_workbook(table, data_only=True, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    hdr = next(rows)
    iM = _idx(hdr, 'Mdt')
    iL = hdr.index('Langue')
    iH = _idx(hdr, 'L.HCM', 'GrPay')
    iR = hdr.index('Rubrique')
    iT = hdr.index('Libellé de rubrique')
    out: dict[str, str] = {}
    for r in rows:
        if not _row_matches(r, iM, cfg["mdt"], iH, cfg["grpay"]):
            continue
        if r[iL] != cfg["lang_fr"]:
            continue
        rub = r[iR]
        if not rub:
            continue
        out[rub] = r[iT] or ''
    wb.close()
    print(f"      → {len(out)} libellés FR trouvés")
    return out


def load_t554e(cfg: dict) -> list[dict] | None:
    """Charge T554E → liste ordonnée [{'clabs': '1', 'label': 'Congé'}, …].

    T554E est le référentiel des classes d'absences (CLABS → libellé). Chaque
    classe devient une colonne du bloc pénalisant dans le WTC, dans l'ordre
    croissant des CLABS (1, 10, 20, 30, 40, 50, 60, 70, 80, 90, 91).

    Filtre : GrPay = cfg['grpay']. On NE filtre PAS sur Mdt : dans les dumps,
    le Mdt de T554E (référentiel) peut différer du Mdt paie.

    Renvoie None si la table n'est pas configurée / introuvable pour le client
    (ex. AKN) → l'appelant retombe sur DEFAULT_PEN_CLASSES (11 classes).
    """
    if "T554E" not in cfg["tables"]:
        return None
    try:
        table = _resolve_table(cfg, "T554E")
    except FileNotFoundError:
        return None
    print(f"[4/5 g] Chargement T554E ({table.name}, GrPay={cfg['grpay']})…")
    wb = openpyxl.load_workbook(table, data_only=True, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    hdr = next(rows)
    iG = _idx(hdr, 'GrPay', 'L.HCM')
    iC = hdr.index('CLABS')
    iL = hdr.index('Classe')
    out: list[dict] = []
    seen: set[str] = set()
    for r in rows:
        if iG is not None and r[iG] != cfg["grpay"]:
            continue
        clabs = r[iC]
        if clabs in (None, ''):
            continue
        clabs = str(clabs)
        if clabs in seen:
            continue
        seen.add(clabs)
        out.append({'clabs': clabs, 'label': r[iL] or clabs})
    wb.close()
    # Ordre croissant des CLABS (1, 10, 20, …, 91) → mappe P, Q, R, … Z
    out.sort(key=lambda d: int(d['clabs']) if d['clabs'].isdigit() else 1_000_000)
    print(f"      → {len(out)} classes d'absences")
    return out


def load_calendar(cfg: dict) -> list[dict]:
    """Charge les codes de l'onglet 'Calendrier' depuis T554S.

    Le Calendrier liste les TYPES d'absence — champ T554S.TypAb (code à 2 lettres :
    AI, MA, CP, RT…) — avec leur libellé (T554S 'Texte cat. prés./abs.'). TypAb
    n'est renseigné par SAP que sur certaines lignes "représentatives" de T554S
    (la même CatAbsP peut avoir plusieurs lignes, dont une seule porte le code) ;
    on ne garde donc QUE les lignes où TypAb ≠ vide.

    Filtre : GrSdP = cfg['grsdp']. Le texte inline de T554S est dans la langue du
    grouping → français pour GrSdP=6 (pas de filtre langue séparé : T554S n'a pas
    de colonne Langue, contrairement à T554T).

    Dédup : 1ʳᵉ occurrence (ordre des lignes du dump). Le CODE est fixé à sa 1ʳᵉ
    occurrence ; le LIBELLÉ prend la 1ʳᵉ occurrence NON VIDE (certaines lignes
    "représentatives" de T554S portent le TypAb mais ont un texte vide, alors
    qu'une ligne suivante du même code en a un — ex. EF). Les codes dont aucune
    occurrence n'a de texte (ex. MD/SF/PT/FT chez ABV) restent sans libellé.
    Tri final alphabétique par code (comme le template de référence).

    Renvoie une liste ordonnée de dicts {'code': 'AI', 'label': 'Absence injustifiée'}.
    """
    table = _resolve_table(cfg, "T554S")
    print(f"[5/5 c] Chargement Calendrier depuis T554S "
          f"({table.name}, GrSdP={cfg['grsdp']})…")
    wb = openpyxl.load_workbook(table, data_only=True, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    hdr = next(rows)
    iG = hdr.index('GrSdP')
    iTy = hdr.index('TypAb')
    iTx = hdr.index('Texte cat. prés./abs.')
    seen: dict[str, str] = {}
    for r in rows:
        if r[iG] != cfg["grsdp"]:
            continue
        code = r[iTy]
        if not (isinstance(code, str) and code.strip()):
            continue
        code = code.strip()
        label = (r[iTx] or '').strip()
        if code not in seen:
            seen[code] = label              # 1ʳᵉ occurrence : fixe le code
        elif not seen[code] and label:
            seen[code] = label              # complète un libellé resté vide
    wb.close()
    out = [{'code': c, 'label': seen[c]} for c in sorted(seen)]
    empties = [c for c in sorted(seen) if not seen[c]]
    print(f"      → {len(out)} codes calendrier (TypAb) extraits")
    if empties:
        print(f"      ⚠ {len(empties)} code(s) sans libellé dans T554S "
              f"(à compléter à la main) : {', '.join(empties)}")
    return out


def _classify_rubrique(rub: str, libelle: str) -> str:
    """Renvoie 'paiement', 'retenue' ou '' selon le libellé T512T.
    L'ordre des sous-blocs dans T554C n'est pas fiable (parfois retenue en
    premier), donc on s'appuie sur la convention de nommage des libellés AKN :
       - 'Paiement…', 'Paiem.…' → paiement
       - 'Retenue…', 'Ret.…', 'Reten…' → retenue
    """
    if not libelle:
        return ''
    low = libelle.lower().lstrip()
    if low.startswith(('paiement', 'paiem.', 'paiem ', 'pay.', 'pay ')):
        return 'paiement'
    if low.startswith(('retenue', 'ret.', 'ret ', 'reten')):
        return 'retenue'
    return ''


def load_t554c_rules(cfg: dict, t512t: dict[str, str]) -> dict[str, dict]:
    """Charge T554C → dict { règle_valorisation: {'paiement': rub, 'retenue': rub} }.

    T554C contient 15 sous-blocs (DH, Pourc., Tp, RB, Rubrique, RègleJourn). Pour
    chaque règle de valorisation on extrait les 2 premières rubriques non vides,
    puis on les classe en Paiement/Retenue via les libellés T512T (les libellés
    AKN commencent par 'Paiement…' ou 'Retenue…' — fiable).
    Si plusieurs lignes existent pour la même règle (versions par dates), on garde
    celle dont la date Fin est la plus haute.
    """
    table = _resolve_table(cfg, "T554C")
    print(f"[4/5 f] Chargement T554C ({table.name}, Grpe={cfg['grpay']})…")
    wb = openpyxl.load_workbook(table, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]
    iM = _idx(hdr, 'Mdt')
    iL = _idx(hdr, 'L.HCM', 'GrPay')
    iG = hdr.index('Grpe')
    iR = hdr.index('Règle de valorisat.')
    iF = hdr.index('Fin')

    # Indices des 15 colonnes "Rubrique" (2e moitié) et des 10 colonnes "CLABS"
    # (1re moitié — sous-blocs de pénalisation).
    rubrique_cols = [i for i, h in enumerate(hdr) if h == 'Rubrique']
    clabs_cols = [i for i, h in enumerate(hdr) if h == 'CLABS']

    by_rule: dict[str, tuple] = {}
    n_swapped = 0
    for r in rows[1:]:
        if not _row_matches(r, iM, cfg["mdt"], iL, cfg["grpay"]):
            continue
        if r[iG] != cfg["grpay"]:
            continue
        rule = r[iR]
        if not rule:
            continue
        fin = r[iF]
        rubriques = [r[i] for i in rubrique_cols if r[i]]
        first = rubriques[0] if rubriques else ''
        second = rubriques[1] if len(rubriques) >= 2 else ''

        # Classification par libellé (T512T) — l'ordre dans T554C n'est pas fiable
        c1 = _classify_rubrique(first, t512t.get(first, ''))
        c2 = _classify_rubrique(second, t512t.get(second, ''))
        paiement, retenue = first, second
        if c1 == 'retenue' and c2 == 'paiement':
            paiement, retenue = second, first
            n_swapped += 1
        elif c1 == 'retenue' and not c2:
            paiement, retenue = '', first
        elif c1 == 'paiement' and c2 == 'paiement':
            # Cas rare — on garde l'ordre
            pass

        # Liste des CLABS pénalisées (ignore '0' qui veut dire "pas pénalisant")
        clabs = {str(r[i]) for i in clabs_cols
                 if r[i] not in (None, '', '0')}

        prev = by_rule.get(rule)
        if prev is None or (fin and (not prev[3] or fin > prev[3])):
            by_rule[rule] = (paiement, retenue, clabs, fin)
    wb.close()
    out = {k: {'paiement': v[0], 'retenue': v[1], 'clabs': v[2]}
           for k, v in by_rule.items()}
    print(f"      → {len(out)} règles de valorisation indexées "
          f"({n_swapped} swaps Paiement/Retenue corrigés)")
    return out


def apply_therapeutic_partial_override(data: list[dict], cfg: dict) -> int:
    """Override métier ABVIE pour les absences "Temps partiel thérapeutique"
    (CatAbsP 1200-1209), piloté par cfg['therapeutic_partial'].

    - Col D (Rubrique Paiement) = texte fixe 'IT 2010' : les rubriques de
      paiement de ces absences sont dans l'infotype 2010 (envoyé par le client).
      HRO refuse d'afficher un n° de rubrique qui laisserait croire à un
      déclenchement automatique → on n'affiche que 'IT 2010' (pas de libellé E
      ni d'unité F, qui n'auraient pas de sens sans rubrique).
    - Col G (Rubrique Retenue) : si vide, recopie la rubrique de retenue du code
      source (1200, généralement 3170) + son libellé (H) et son unité (I). Cette
      retenue se déclenche automatiquement mais n'est pas indiquée dans le dump.

    Sans clé 'therapeutic_partial' dans cfg (ex. AKN), ne fait rien. Renvoie le
    nombre de codes modifiés.
    """
    ov = cfg.get("therapeutic_partial")
    if not ov:
        return 0
    codes = set(ov["codes"])
    src_code = ov["retenue_source"]
    pay_label = ov["payment_label"]
    src = next((r for r in data if r['CatAbsP'] == src_code), None)
    src_retenue = ((src['RubRetenue'], src['LibRetenue'], src['UnitRetenue'])
                   if src else ('', '', ''))
    n = 0
    for rec in data:
        if rec['CatAbsP'] not in codes:
            continue
        # Col D : texte fixe (pas de rubrique → on neutralise libellé/unité paie)
        rec['RubPaiement'] = pay_label
        rec['LibPaiement'] = ''
        rec['UnitPaiement'] = ''
        # Col G : recopie de la retenue du code source si la retenue est vide
        if not rec['RubRetenue'] and rec['CatAbsP'] != src_code:
            rec['RubRetenue'], rec['LibRetenue'], rec['UnitRetenue'] = src_retenue
        n += 1
    return n


def load_client_data(cfg: dict) -> list[dict]:
    """Charge les catégories d'absences pour le client depuis T554S + T554T.

    Filtre :
      - T554S : GrSdP = cfg['grsdp']
      - T554T : GrSdP = cfg['grsdp'] ET Langue = cfg['lang_fr']

    Renvoie une liste de dicts ordonnée par CatAbsP.
    """
    table_s = _resolve_table(cfg, "T554S")
    print(f"[4/5 a] Chargement T554S ({table_s.name}, GrSdP={cfg['grsdp']})…")
    wb554s = openpyxl.load_workbook(table_s, data_only=True, read_only=False)
    ws = wb554s.active
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]
    iS = {name: hdr.index(name) for name in
          ('GrSdP', 'CatAbsP', 'Règle de valorisat.', 'Classe', 'TypAb', 'Fin', 'Début')}
    # Une CatAbsP peut avoir plusieurs lignes (versions par dates). On garde
    # celle dont la date Fin est la plus haute (la plus récente / 9999-12-31).
    by_cat = {}
    for r in rows[1:]:
        if r[iS['GrSdP']] != cfg["grsdp"]:
            continue
        cat = r[iS['CatAbsP']]
        if not cat:
            continue
        fin = r[iS['Fin']]
        prev = by_cat.get(cat)
        if prev is None or (fin and (not prev['_fin'] or fin > prev['_fin'])):
            by_cat[cat] = {
                'CatAbsP': cat,
                'RegleValorisat': r[iS['Règle de valorisat.']] or '',
                'Classe': r[iS['Classe']] or '',
                'TypAb': r[iS['TypAb']] or '',
                '_fin': fin,
            }
    print(f"      → {len(by_cat)} catégories trouvées")

    table_t = _resolve_table(cfg, "T554T")
    print(f"[4/5 b] Chargement T554T ({table_t.name}, Langue={cfg['lang_fr']}, "
          f"GrSdP={cfg['grsdp']})…")
    wb554t = openpyxl.load_workbook(table_t, data_only=True, read_only=False)
    ws2 = wb554t.active
    rows2 = list(ws2.iter_rows(values_only=True))
    hdr2 = rows2[0]
    iT = {name: hdr2.index(name) for name in
          ('Langue', 'GrSdP', 'CatAbsP', 'Texte cat. prés./abs.')}
    labels = {}
    for r in rows2[1:]:
        if r[iT['Langue']] != cfg["lang_fr"] or r[iT['GrSdP']] != cfg["grsdp"]:
            continue
        labels[r[iT['CatAbsP']]] = r[iT['Texte cat. prés./abs.']]
    print(f"      → {len(labels)} libellés FR trouvés")

    # Charge les ressources nécessaires pour enrichir les colonnes D-I
    units_map = load_units_map()
    t511 = load_t511(cfg)
    t512t = load_t512t(cfg)
    t554c = load_t554c_rules(cfg, t512t)

    def lookup_unit(rub: str) -> str:
        """rubrique → libellé d'unité humain via T511 + numeros_vs_unités."""
        if not rub:
            return ''
        info = t511.get(rub)
        unt = info['unt'] if info else ''
        return units_map.get(unt, unt) if unt else ''

    def lookup_nombre_montant(cat: str, rub_p: str, rub_r: str) -> tuple[str, str]:
        """Détermine les textes (Nombre, Montant) à écrire en cols J/K du WTC.

        Ordre de recherche dans T511.C (Rubrique) :
          1. rub_p (col D = Rubrique Paiement) — vraie rubrique de paie
          2. rub_r (col G = Rubrique Retenue)  — vraie rubrique de paie
          3. cat   (col A = CatAbsP)           — fallback (peut collisionner par
             hasard avec une rubrique de paie sans rapport)
        Le 1er candidat trouvé dans T511 donne le caractère 'c' qui sert de
        clé dans NOMBRE_MONTANT_MAP. Si le 'c' trouvé est vide (espace), on
        laisse J et K du WTC vides (cas fréquent sur rubriques 9000+).
        Si aucun candidat n'est trouvé dans T511, J et K restent aussi vides.
        """
        for candidate in (rub_p, rub_r, cat):
            if not candidate:
                continue
            info = t511.get(candidate)
            if info is None:
                continue
            c = info['c']
            if not c or c == ' ':
                return ('', '')
            return NOMBRE_MONTANT_MAP.get(c, ('', ''))
        return ('', '')

    # Combine et tri
    out = []
    n_paiement = 0
    n_retenue = 0
    n_jk = 0
    for cat in sorted(by_cat.keys()):
        rec = by_cat[cat]
        rule = rec['RegleValorisat']
        rubs = t554c.get(rule, {'paiement': '', 'retenue': '', 'clabs': set()})
        rub_p = rubs['paiement']
        rub_r = rubs['retenue']
        clabs = rubs.get('clabs', set())
        if rub_p:
            n_paiement += 1
        if rub_r:
            n_retenue += 1
        jk_nombre, jk_montant = lookup_nombre_montant(cat, rub_p, rub_r)
        if jk_nombre or jk_montant:
            n_jk += 1
        out.append({
            'CatAbsP': cat,
            'Libelle': labels.get(cat, '(libellé FR manquant)'),
            'RegleValorisat': rule,
            'Classe': rec['Classe'],
            'TypAb': rec['TypAb'],
            'Categorie': cat[:2] + 'xx' if len(cat) >= 2 else cat,
            # Enrichissement v0.4.0
            'RubPaiement': rub_p,
            'LibPaiement': t512t.get(rub_p, '') if rub_p else '',
            'UnitPaiement': lookup_unit(rub_p),
            'RubRetenue': rub_r,
            'LibRetenue': t512t.get(rub_r, '') if rub_r else '',
            'UnitRetenue': lookup_unit(rub_r),
            # Enrichissement v0.6.0 — cols J + K via T511.C
            'Nombre': jk_nombre,
            'Montant': jk_montant,
            # Enrichissement v0.8.0 — col L (Pilotage Paie : unité de temps)
            # Basée sur rub Paiement (D) prioritaire, sinon rub Retenue (G).
            'UnitL': lookup_unit(rub_p) or lookup_unit(rub_r),
            # Enrichissement v1.0.0 — ensemble brut des CLABS pénalisées de la
            # règle de valorisation. Le mapping CLABS → colonne (P..Z) est fait
            # dans populate_main_sheet à partir des classes T554E.
            'ClabsSet': clabs,
        })
    print(f"      → {n_paiement} rub. Paiement / {n_retenue} rub. Retenue résolues")
    print(f"      → {n_jk} lignes avec textes Nombre/Montant remplis depuis T511.C")

    # Override métier ABVIE — "Temps partiel thérapeutique" (D='IT 2010', G recopié de 1200)
    n_ov = apply_therapeutic_partial_override(out, cfg)
    if n_ov:
        print(f"      → override 'Temps partiel théra' appliqué sur {n_ov} codes "
              f"(D='IT 2010', G recopié du code source)")
    return out


def _safe_sheet_name(label: str, suffix: str = "_Data") -> str:
    """Nom d'onglet Excel valide : retire les caractères interdits (:\\/?*[])
    et borne la longueur à 31 caractères (suffixe inclus)."""
    cleaned = ''.join('_' if ch in r':\/?*[]' else ch for ch in (label or 'WTC')).strip()
    cleaned = cleaned or 'WTC'
    return cleaned[:31 - len(suffix)] + suffix


def add_data_sheet(wb, data: list[dict], cfg: dict) -> None:
    """Peuple l'onglet de données à plat 'Wagetype Catalog Absence_Data' avec les
    catégories du client. L'onglet du template est réutilisé EN PLACE (sa position
    — 3e onglet, juste après Parameter (masqué) — est conservée) ; on le vide
    d'abord de ses données d'exemple. S'il n'existe pas (template inhabituel), on
    le crée."""
    client = cfg.get("label", "WTC")
    sheet_name = DATA_SHEET
    print(f"[5/5 a] Peuplement de l'onglet '{sheet_name}' ({len(data)} lignes)…")
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Vider les données d'exemple en conservant la position de l'onglet.
        for mr in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(mr))
        if ws.max_row >= 1:
            ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(sheet_name)

    from openpyxl.styles import Alignment, Border, Side

    # Titre
    ws['A1'] = f"Données {client} — Catégories d'absences/présences"
    ws['A1'].font = Font(name='Arial', size=14, bold=True, color=WHITE)
    ws['A1'].fill = PatternFill('solid', fgColor=STRADA_DARK)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:F1')
    ws.row_dimensions[1].height = 28

    ws['A2'] = (f"Source : T554S + T554T  |  GrSdP={cfg['grsdp']}  |  "
                f"Langue={cfg['lang_fr']}  |  GrPay={cfg['grpay']}  |  "
                f"Customer Code={client}")
    ws['A2'].font = Font(name='Arial', size=9, italic=True, color="FF333333")
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:F2')

    # Entêtes
    headers = ['Catégorie', 'Code (CatAbsP)', 'Libellé Absence (FR)',
               'Règle Valorisat.', 'Classe', 'TypAb']
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col_idx, value=h)
        c.font = Font(name='Arial', size=10, bold=True, color=WHITE)
        c.fill = PatternFill('solid', fgColor=STRADA_MED)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[4].height = 32

    # Données
    thin = Side(border_style="thin", color="FF888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    prev_cat = None
    for i, rec in enumerate(data, start=5):
        # Séparateur visuel quand la catégorie change
        cat_changed = prev_cat is not None and rec['Categorie'] != prev_cat
        row_fill = STRADA_MED if cat_changed else YELLOW

        for col_idx, key in enumerate(
            ['Categorie', 'CatAbsP', 'Libelle', 'RegleValorisat', 'Classe', 'TypAb'],
            start=1,
        ):
            c = ws.cell(row=i, column=col_idx, value=rec[key])
            c.font = Font(name='Arial', size=9,
                          bold=cat_changed,
                          color=WHITE if cat_changed else "FF000000")
            c.fill = PatternFill('solid', fgColor=row_fill)
            c.border = border
            c.alignment = Alignment(
                horizontal='center' if col_idx != 3 else 'left',
                vertical='center',
            )
        prev_cat = rec['Categorie']

    # Largeurs de colonnes
    widths = {'A': 12, 'B': 14, 'C': 40, 'D': 16, 'E': 9, 'F': 9}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = 'A5'
    print(f"      → onglet {sheet_name} peuplé ({len(data)} lignes)")


def populate_calendar_sheet(wb, calendar: list[dict]) -> int:
    """Peuple l'onglet 'Calendrier' avec les types d'absence (T554S.TypAb).

    Réécrit l'onglet EN PLACE : la ligne d'en-tête (1) est conservée, toutes les
    lignes en dessous sont supprimées puis réécrites — d'abord les codes
    (col A = code, col B = libellé), puis le bloc vert clair 'Entrées spécifiques'
    (zone de saisie manuelle de l'HRO) réinséré juste sous les données.

    Les styles (police/bordure/alignement des cellules de données, fond vert clair
    du bloc 'Entrées spécifiques') sont capturés sur le template puis ré-appliqués,
    pour que le rendu reste identique quel que soit le nombre de codes du client.

    Renvoie le nombre de codes écrits."""
    if CALENDAR_SHEET not in wb.sheetnames:
        print(f"      ⚠ onglet '{CALENDAR_SHEET}' absent du template — ignoré")
        return 0
    ws = wb[CALENDAR_SHEET]
    print(f"[5/5 d] Peuplement de l'onglet '{CALENDAR_SHEET}' "
          f"({len(calendar)} codes)…")

    # Styles de référence capturés AVANT nettoyage (A2/B2 = données, A32 = bloc vert,
    # E34 = libellé flottant 'Entrées spécifiques').
    code_font, code_border, code_align = (
        copy(ws['A2'].font), copy(ws['A2'].border), copy(ws['A2'].alignment))
    label_font, label_border, label_align = (
        copy(ws['B2'].font), copy(ws['B2'].border), copy(ws['B2'].alignment))
    green_font, green_fill, green_border = (
        copy(ws['A32'].font), copy(ws['A32'].fill), copy(ws['A32'].border))
    extra_val = ws['E34'].value
    extra_font = copy(ws['E34'].font)

    # Nettoyage : supprimer toutes les lignes sous l'en-tête.
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    # Données (déjà triées alpha par load_calendar).
    for i, rec in enumerate(calendar, start=2):
        a = ws.cell(row=i, column=1, value=rec['code'])
        a.font, a.border, a.alignment = copy(code_font), copy(code_border), copy(code_align)
        b = ws.cell(row=i, column=2, value=rec['label'])
        b.font, b.border, b.alignment = copy(label_font), copy(label_border), copy(label_align)

    # Bloc 'Entrées spécifiques' (vert clair) sous les données.
    start = 2 + len(calendar)
    for r in range(start, start + CAL_EXTRA_ROWS):
        for c in (1, 2):
            cell = ws.cell(row=r, column=c, value=None)
            cell.font, cell.fill, cell.border = copy(green_font), copy(green_fill), copy(green_border)
    # Libellé flottant en col E, au milieu du bloc.
    lbl = ws.cell(row=start + CAL_EXTRA_ROWS // 2, column=5, value=extra_val)
    lbl.font = copy(extra_font)

    print(f"      → onglet {CALENDAR_SHEET} peuplé ({len(calendar)} codes)")
    return len(calendar)


# ---- Listes déroulantes (data validations) -----------------------------------

def _dv_kind(formula1: str) -> str | None:
    """Identifie le 'genre' d'une data validation d'après sa formule source.
    Renvoie une clé interne ('us', 'unit', 'L', 'th', 'tj') ou None."""
    f = (formula1 or '').strip()
    if f == 'Us':
        return 'us'
    if f == 'TH':
        return 'th'
    if f == 'TJ':
        return 'tj'
    if 'Jours calendrier' in f:
        return 'unit'
    if f.strip('"') == 'Heure,Jour':
        return 'L'
    return None


def _broken_named_ranges(wb) -> set[str]:
    """Renvoie l'ensemble des named ranges dont la cible est cassée (#REF!).
    Dans WTCA reference, TH et TJ sont #REF! (l'onglet 'Taux de valorisation'
    qui les définissait a été supprimé) → on ne ré-applique pas ces listes."""
    broken: set[str] = set()
    try:
        items = list(wb.defined_names.items())
    except AttributeError:           # openpyxl < 3.1
        items = [(n, wb.defined_names[n]) for n in wb.defined_names]
    for name, dn in items:
        if dn is not None and '#REF!' in str(getattr(dn, 'value', '')):
            broken.add(name)
    return broken


def capture_data_validations(ws) -> dict[str, tuple]:
    """Photographie les définitions de listes déroulantes du template, indexées
    par genre (_dv_kind). Renvoie {kind: (type, formula1, allow_blank)}."""
    defs: dict[str, tuple] = {}
    for dv in ws.data_validations.dataValidation:
        kind = _dv_kind(dv.formula1)
        if kind and kind not in defs:
            defs[kind] = (dv.type, dv.formula1, dv.allow_blank)
    return defs


def _col_ranges(col: int, rows: list[int]) -> str:
    """Compacte une liste de lignes en plages Excel pour une colonne donnée :
    col=3, rows=[9,10,11,13] → 'C9:C11 C13:C13'."""
    if not rows:
        return ''
    L = get_column_letter(col)
    rows = sorted(rows)
    out = []
    i = 0
    while i < len(rows):
        j = i
        while j + 1 < len(rows) and rows[j + 1] == rows[j] + 1:
            j += 1
        out.append(f"{L}{rows[i]}:{L}{rows[j]}")
        i = j + 1
    return ' '.join(out)


def reapply_data_validations(ws, defs: dict[str, tuple], data_rows: list[int],
                             n_pen: int, broken: set[str]) -> int:
    """Ré-applique les listes déroulantes du template aux vraies lignes de données.

    Le template définit les validations sur les lignes d'exemple (9-156) ; après
    réécriture des données, on remet chaque liste sur les lignes réellement
    occupées (data_rows, hors entêtes de catégorie). Mapping colonnes → genre :
       - 'us'  → C (Used) + P..Z (classes d'absences) : liste ■/□
       - 'unit'→ F + I (unité de temps Paiement / Retenue)
       - 'L'   → L (unité de temps Pilotage Paie : Heure/Jour)
       - 'th'  → N (Taux horaire)   ← IGNORÉ si #REF!
       - 'tj'  → O (Taux jour)      ← IGNORÉ si #REF!
    Renvoie le nombre de validations ré-appliquées.
    """
    from openpyxl.worksheet.datavalidation import DataValidation

    # Colonnes 'us' : Used (C=3) + bloc des classes (P.. = 16..15+n_pen).
    # NB : on n'applique PAS ■/□ aux colonnes J/K (Nombre/Montant) — le template
    # le faisait mais leur contenu est du texte ('oblig.', 'facult.'…), pas ■/□.
    groups = {
        'us':   [3] + list(range(PEN_START_COL, PEN_START_COL + n_pen)),
        'unit': [6, 9],
        'L':    [12],
        'th':   [14],
        'tj':   [15],
    }
    ws.data_validations.dataValidation = []
    applied = 0
    for kind, cols in groups.items():
        d = defs.get(kind)
        if not d:
            continue
        dtype, formula1, allow_blank = d
        if (formula1 or '').strip() in broken:
            continue  # source #REF! (TH/TJ) → on n'ajoute pas de liste cassée
        sqref = ' '.join(filter(None, (_col_ranges(c, data_rows) for c in cols)))
        if not sqref:
            continue
        dv = DataValidation(type=dtype, formula1=formula1, allow_blank=allow_blank)
        dv.sqref = sqref
        ws.add_data_validation(dv)
        applied += 1
    return applied


def restructure_penalisant_block(ws, pen_classes: list[dict]) -> int:
    """Étend le bloc des classes d'absences de PEN_TEMPLATE_COUNT colonnes (P-S
    dans le template GFK) à len(pen_classes) colonnes, en DÉCALANT vers la droite
    les colonnes "trailing" (Additional Comments, WD absence name, Customizing…)
    au lieu de les écraser, puis réécrit les entêtes ligne 7 avec les libellés
    'Classe' de T554E.

    shift = len(pen_classes) − PEN_TEMPLATE_COUNT. Si shift == 0 (client sans
    T554E → 4 pénalisants), aucune colonne n'est déplacée : on se contente de
    réécrire les 4 entêtes. Renvoie le shift appliqué.
    """
    n = len(pen_classes)
    shift = n - PEN_TEMPLATE_COUNT
    pen_start = PEN_START_COL                          # 16 = P
    trailing_start = pen_start + PEN_TEMPLATE_COUNT    # 20 = T (1re col à décaler)
    trailing_end = TRAILING_END_COL                    # 25 = Y
    pilotage_end = pen_start + PEN_TEMPLATE_COUNT - 1   # 19 = S (fin merge J6:S6)

    if shift > 0:
        # 1. Snapshot des entêtes trailing (lignes 6-7) : valeur, styles, largeur.
        snap = []  # (old_col, width, {row: (value, font, fill, border, align, numfmt)})
        for col in range(trailing_start, trailing_end + 1):
            L = get_column_letter(col)
            dim = ws.column_dimensions.get(L)
            cells = {}
            for r in (6, 7):
                c = ws.cell(row=r, column=col)
                cells[r] = (c.value, copy(c.font), copy(c.fill),
                            copy(c.border), copy(c.alignment), c.number_format)
            snap.append((col, dim.width if dim else None, cells))

        # 2. Repérer les fusions d'entête concernées :
        #    - "Pilotage Paie" J6:S6 (cols 10..pilotage_end) → à étendre
        #    - fusions entièrement dans la zone trailing (ex. V6:Y6) → à décaler
        pilotage = None
        shifted_merges = []
        for mr in list(ws.merged_cells.ranges):
            if mr.min_row > HEADER_ROW_END:
                continue
            if mr.min_col == 10 and mr.max_col == pilotage_end:
                pilotage = (mr.min_row, mr.min_col, mr.max_row, mr.max_col)
            elif mr.min_col >= trailing_start and mr.max_col <= trailing_end:
                shifted_merges.append((mr.min_row, mr.min_col, mr.max_row, mr.max_col))
        if pilotage is not None:
            ws.unmerge_cells(start_row=pilotage[0], start_column=pilotage[1],
                             end_row=pilotage[2], end_column=pilotage[3])
        for (r0, c0, r1, c1) in shifted_merges:
            ws.unmerge_cells(start_row=r0, start_column=c0, end_row=r1, end_column=c1)

        # 3. Effacer les anciennes cellules trailing (lignes 6-7).
        for col in range(trailing_start, trailing_end + 1):
            for r in (6, 7):
                ws.cell(row=r, column=col).value = None

        # 4. Réécrire le contenu trailing décalé de +shift (valeurs/styles/largeurs).
        for (old_col, width, cells) in snap:
            new_col = old_col + shift
            if width is not None:
                ws.column_dimensions[get_column_letter(new_col)].width = width
            for r, (val, font, fill, border, align, numfmt) in cells.items():
                nc = ws.cell(row=r, column=new_col, value=val)
                nc.font = font
                nc.fill = fill
                nc.border = border
                nc.alignment = align
                nc.number_format = numfmt

        # 5. Recréer les fusions trailing décalées (V6:Y6 → AC6:AF6).
        for (r0, c0, r1, c1) in shifted_merges:
            ws.merge_cells(start_row=r0, start_column=c0 + shift,
                           end_row=r1, end_column=c1 + shift)

        # 6. Étendre "Pilotage Paie" pour englober tout le bloc de classes.
        if pilotage is not None:
            ws.merge_cells(start_row=6, start_column=10,
                           end_row=6, end_column=pen_start + n - 1)

    # 7. (Ré)écrire les entêtes ligne 7 du bloc, libellés = 'Classe' T554E.
    #    Style de référence = ancien entête pénalisant (S7 : orange, gras, centré).
    ref = ws.cell(row=7, column=pilotage_end)          # S7
    ref_style = (copy(ref.font), copy(ref.fill), copy(ref.border), copy(ref.alignment))
    ref_dim = ws.column_dimensions.get(get_column_letter(pilotage_end))
    ref_width = ref_dim.width if ref_dim else None
    for i, pc in enumerate(pen_classes):
        col = pen_start + i
        c = ws.cell(row=7, column=col, value=pc['label'])
        c.font, c.fill, c.border, c.alignment = (copy(s) for s in ref_style)
        ws.column_dimensions[get_column_letter(col)].width = ref_width or 11.55
    return shift


def populate_main_sheet(wb, data: list[dict], cfg: dict,
                        pen_classes: list[dict]) -> int:
    """Remplace les données du sheet principal 'Absences-Présences' par
    les vraies données client, groupées par catégorie '(0Xxx)'."""
    from openpyxl.styles import Alignment, Border, Side

    print(f"[5/5 b] Injection données dans onglet '{MAIN_SHEET}'…")
    ws = wb[MAIN_SHEET]

    # Photographier les listes déroulantes du template AVANT d'écrire (les ranges
    # du template visent les lignes d'exemple ; on les remettra sur les vraies
    # lignes en fin de fonction). On note aussi les named ranges cassés (#REF!).
    dv_defs = capture_data_validations(ws)
    broken_names = _broken_named_ranges(wb)

    # 0. Restructurer le bloc des classes d'absences (P..) selon T554E, en
    #    décalant les colonnes "trailing" vers la droite. Renvoie le décalage.
    shift = restructure_penalisant_block(ws, pen_classes)
    n_pen = len(pen_classes)
    pen_end = PEN_START_COL + n_pen - 1            # ex. 26 = Z pour 11 classes
    # Largeur utile totale = anciennes colonnes (jusqu'à Y=25) + décalage.
    full_width = TRAILING_END_COL + shift          # ex. 32 = AF pour shift=7
    print(f"      → {n_pen} classes d'absences (cols {get_column_letter(PEN_START_COL)}"
          f"-{get_column_letter(pen_end)}), trailing décalé de +{shift}")

    # 1. Mettre à jour le titre (ex. "FR Absences" → "AKN Absences")
    old, new = cfg["title_replace"]
    for r in range(1, HEADER_ROW_END + 1):
        for c in range(1, full_width + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value and isinstance(cell.value, str) and "Absences/Présences Catalogue" in cell.value:
                cell.value = cell.value.replace(old, new)

    # 2. Supprimer toutes les fusions dans la zone de données
    to_remove = [mr for mr in list(ws.merged_cells.ranges)
                 if mr.min_row >= DATA_ROW_START]
    for mr in to_remove:
        ws.unmerge_cells(str(mr))

    # 3. Vider valeurs + reset basique des styles dans les lignes 8-156
    blank_fill = PatternFill(fill_type=None)
    blank_font = Font(name='Arial', size=10, color=BLACK)
    thin = Side(border_style="thin", color="FF888888")
    blank_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(DATA_ROW_START, DATA_ROW_END + 1):
        for c in range(1, full_width + 1):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.fill = blank_fill
            cell.font = blank_font

    # 4. Grouper les données par catégorie (préfixe '0Xxx', '1Hxx', 'P1xx'…)
    by_category: dict[str, list[dict]] = {}
    cat_order: list[str] = []
    for rec in data:
        cat = rec['Categorie']
        if cat not in by_category:
            by_category[cat] = []
            cat_order.append(cat)
        by_category[cat].append(rec)

    # 5. Écrire les données par catégorie
    cat_font = Font(name='Arial', size=10, bold=True, color=WHITE)
    cat_fill = PatternFill('solid', fgColor=STRADA_MED)
    cat_align = Alignment(horizontal='left', vertical='center', indent=1)
    data_font = Font(name='Arial', size=10, color=BLACK)
    data_fill = PatternFill('solid', fgColor=YELLOW)
    code_align = Alignment(horizontal='center', vertical='center')
    label_align = Alignment(horizontal='left', vertical='center', indent=1)
    used_font = Font(name='Arial', size=14, color="FF666666")
    used_align = Alignment(horizontal='center', vertical='center')

    cur_row = DATA_ROW_START
    n_data_rows = 0
    data_rows: list[int] = []   # lignes réellement occupées (hors entêtes catégorie)
    for cat in cat_order:
        recs = by_category[cat]
        # Heuristique du nom : libellé du 1er code dont le libellé existe
        cat_name = next((r['Libelle'] for r in recs
                         if r['Libelle'] and 'manquant' not in r['Libelle']), '')
        # Garder seulement les 3-4 premiers mots pour un nom court
        cat_short = ' '.join(cat_name.split()[:3]) if cat_name else ''
        cat_title = f"({cat}) {cat_short}".strip()

        # Ligne d'entête de catégorie : merge A:B, fond vert moyen, texte blanc gras
        ws.cell(row=cur_row, column=1, value=cat_title).font = cat_font
        ws.cell(row=cur_row, column=1).fill = cat_fill
        ws.cell(row=cur_row, column=1).alignment = cat_align
        for c in range(2, full_width + 1):
            ws.cell(row=cur_row, column=c).fill = cat_fill
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=2)
        ws.row_dimensions[cur_row].height = 22
        cur_row += 1

        # Lignes de données
        for rec in recs:
            # Col A : Code
            c_a = ws.cell(row=cur_row, column=1, value=rec['CatAbsP'])
            c_a.font = data_font
            c_a.fill = data_fill
            c_a.alignment = code_align
            c_a.border = blank_border
            # Col B : Libellé
            c_b = ws.cell(row=cur_row, column=2, value=rec['Libelle'])
            c_b.font = data_font
            c_b.fill = data_fill
            c_b.alignment = label_align
            c_b.border = blank_border
            # Col C : Used (□ par défaut)
            c_c = ws.cell(row=cur_row, column=3, value='□')
            c_c.font = used_font
            c_c.fill = data_fill
            c_c.alignment = used_align
            c_c.border = blank_border
            # Cols D-I : Paiement (D, E, F) puis Retenue (G, H, I)
            #   D = Rubrique Paiement / E = Libellé rubrique / F = Unité de temps
            #   G = Rubrique Retenue  / H = Libellé rubrique / I = Unité de temps
            paiement_retenue_vals = [
                (4, rec['RubPaiement'], code_align),
                (5, rec['LibPaiement'], label_align),
                (6, rec['UnitPaiement'], code_align),
                (7, rec['RubRetenue'], code_align),
                (8, rec['LibRetenue'], label_align),
                (9, rec['UnitRetenue'], code_align),
            ]
            for col_idx, val, align in paiement_retenue_vals:
                cx = ws.cell(row=cur_row, column=col_idx, value=val or None)
                cx.font = data_font
                cx.fill = data_fill
                cx.alignment = align
                cx.border = blank_border
            # Col J (10) = Nombre / Col K (11) = Montant — issus de T511.C
            for col_idx, val in ((10, rec['Nombre']), (11, rec['Montant'])):
                cj = ws.cell(row=cur_row, column=col_idx, value=val or None)
                cj.font = data_font
                cj.fill = data_fill
                cj.alignment = code_align
                cj.border = blank_border
            # Col L (12) = Unité de temps (Pilotage Paie) — issue de T511.UnT
            # via rub Paiement (D) avec fallback rub Retenue (G)
            cl = ws.cell(row=cur_row, column=12, value=rec['UnitL'] or None)
            cl.font = data_font
            cl.fill = data_fill
            cl.alignment = code_align
            cl.border = blank_border
            # Colonnes M-O : laissées vides (métier) avec fond jaune cohérent
            for col in range(13, 16):
                cd = ws.cell(row=cur_row, column=col)
                cd.fill = data_fill
                cd.border = blank_border
            # Cols P.. = classes d'absences (T554E) : ■ si la CLABS de la classe
            # figure dans la liste CLABS de la règle de valorisation, □ sinon.
            clabs_set = rec['ClabsSet']
            for i, pc in enumerate(pen_classes):
                col_idx = PEN_START_COL + i
                on = pc['clabs'] in clabs_set
                cp = ws.cell(row=cur_row, column=col_idx, value='■' if on else '□')
                cp.font = used_font
                cp.fill = data_fill
                cp.alignment = used_align
                cp.border = blank_border
            ws.row_dimensions[cur_row].height = 16
            data_rows.append(cur_row)
            cur_row += 1
            n_data_rows += 1

        # Ligne séparatrice "GAP with euHReka standard" si on a marge
        # (style cohérent avec le GFK original — fond jaune, texte vert foncé gras)
        if cur_row <= DATA_ROW_END:
            gap_cell = ws.cell(row=cur_row, column=1, value="(à compléter manuellement)")
            gap_cell.font = Font(name='Arial', size=9, italic=True, color=STRADA_DARK)
            gap_cell.fill = data_fill
            gap_cell.alignment = Alignment(horizontal='center', vertical='center')
            for c in range(2, full_width + 1):
                ws.cell(row=cur_row, column=c).fill = data_fill
            ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=15)
            ws.row_dimensions[cur_row].height = 14
            cur_row += 1

    # Rendre TOUTES les lignes de l'onglet visibles. Le template GFK hérite de
    # ~93 lignes masquées (plan/groupement Excel HRO du catalogue GFK d'origine,
    # lignes 12-156). On ne reconduit pas ce masquage : il cache des codes
    # d'absence valides et ne génère que des questions inutiles côté client. Si
    # HRO ou le client veut masquer des codes non utilisés, ce sera à eux de le
    # faire. On démasque donc toutes les lignes (pas seulement celles remplies,
    # pour couvrir aussi les clients aux données plus courtes type AKN).
    n_unhidden = 0
    for rd in ws.row_dimensions.values():
        if rd.hidden:
            rd.hidden = False
            n_unhidden += 1

    print(f"      → {len(cat_order)} catégories, {n_data_rows} lignes d'absence injectées"
          f" ({n_unhidden} lignes du template démasquées)")

    # Ré-appliquer les listes déroulantes sur les vraies lignes de données.
    n_dv = reapply_data_validations(ws, dv_defs, data_rows, n_pen, broken_names)
    skipped = sorted(broken_names & {'TH', 'TJ'})
    msg = f"      → {n_dv} listes déroulantes ré-appliquées sur {len(data_rows)} lignes"
    if skipped:
        msg += f" (ignorées car #REF! : {', '.join(skipped)})"
    print(msg)
    return n_data_rows


# ---- Interface dossier (HRO) : config dynamique depuis un répertoire ----------

# Tables SAP attendues par l'interface dossier.
REQUIRED_DIR_TABLES = ["T554S", "T554T", "T554C", "T511", "T512T"]
RECOMMENDED_DIR_TABLES = ["T554E"]          # sinon repli DEFAULT_PEN_CLASSES
OPTIONAL_DIR_TABLES = ["Y00BA"]             # non utilisée aujourd'hui
# T508A retirée (v1.4.0) : table des règles de plan de roulement, AUCUN rapport
# avec le Calendrier (généré depuis T554S.TypAb). N'est plus ni scannée ni requise.


def scan_tables_dir(input_dir) -> dict[str, str]:
    """Repère dans input_dir le fichier .xlsx/.xls de chaque table SAP connue.
    Match insensible à la casse sur le nom de fichier (stem == nom de table ;
    Y00BA matche tout fichier commençant par 'y00ba'). Renvoie {table: fichier}."""
    input_dir = Path(input_dir)
    found: dict[str, str] = {}
    if not input_dir.is_dir():
        return found
    known = REQUIRED_DIR_TABLES + RECOMMENDED_DIR_TABLES + OPTIONAL_DIR_TABLES
    for f in sorted(input_dir.iterdir()):
        if not f.is_file() or f.suffix.lower() not in ('.xlsx', '.xls'):
            continue
        stem = f.stem.lower()
        for t in known:
            if t in found:
                continue
            if t == "Y00BA":
                if stem.startswith("y00ba"):
                    found[t] = f.name
            elif stem == t.lower():
                found[t] = f.name
    return found


def build_dir_config(input_dir, output_path, *, revise_previous: bool = False) -> dict:
    """Construit une config de génération depuis un dossier de tables choisi par
    l'utilisateur (interface HRO). Paramètres France par défaut (GrSdP=6,
    GrPay=06, langue=F, pas de filtre Mdt). Tables détectées automatiquement
    (casse libre) via scan_tables_dir().

    revise_previous : réservé (case 'fichier WTC antérieur à réviser') — sans
    effet pour le moment.
    """
    input_dir = Path(input_dir)
    output_path = Path(output_path)
    return {
        "dir": input_dir,
        "tables": scan_tables_dir(input_dir),
        "grsdp": "6",
        "lang_fr": "F",
        "grpay": "06",
        "mdt": None,
        "output": output_path,
        "label": output_path.stem,
        "title_replace": ("FR Absences", "FR Absences"),  # no-op générique
        # Override standard "Temps partiel thérapeutique" (codes SAP 1200-1209).
        "therapeutic_partial": {
            "codes": [str(c) for c in range(1200, 1210)],
            "retenue_source": "1200",
            "payment_label": "IT 2010",
        },
        "revise_previous": revise_previous,  # réservé — sans effet (à venir)
    }


def generate_from_config(cfg: dict) -> dict:
    """Génère le Wagetype Catalog Absence depuis une config (CLIENT_CONFIGS ou
    build_dir_config). Renvoie un dict de statistiques. Les messages de
    progression sont imprimés sur stdout (capturés par la GUI)."""
    output = Path(cfg["output"])
    label = cfg.get("label", "WTC")
    print(f"=== generate_wtc_absences.py v{APP_VERSION} — {label} ===")
    print(f"Template : {WTCA_TEMPLATE.name}")
    print(f"Cible    : {output.name}")

    # Garde-fou : ne jamais écraser le template de référence lui-même.
    if output.resolve() == WTCA_TEMPLATE.resolve():
        raise ValueError(
            f"Le fichier de sortie est le template de référence "
            f"({WTCA_TEMPLATE.name}) — choisissez un autre nom de sortie.")

    # 1. Template Strada déjà finalisé (logo + couleurs + 11 classes en place).
    #    Plus aucun rebranding visuel à faire (cf. v1.3.0).
    if not WTCA_TEMPLATE.exists():
        raise FileNotFoundError(f"Template introuvable : {WTCA_TEMPLATE}")
    print(f"[Template] Chargement {WTCA_TEMPLATE.name}…")
    wb = openpyxl.load_workbook(WTCA_TEMPLATE)
    # 2. Données client (T554S + T554T + enrichissement T554C/T512T/T511)
    data = load_client_data(cfg)
    # 2b. Classes d'absences (T554E) ; repli 11 classes standard si T554E absente
    pen_classes = load_t554e(cfg) or DEFAULT_PEN_CLASSES
    # 2c. Codes calendrier (types d'absence) depuis T554S.TypAb
    calendar = load_calendar(cfg)
    # 3. Onglet de données à plat 'Wagetype Catalog Absence_Data' (peuplé en place)
    add_data_sheet(wb, data, cfg)
    # 4. Injection dans l'onglet principal + ré-application des listes déroulantes
    n_inj = populate_main_sheet(wb, data, cfg, pen_classes)
    # 4b. Onglet 'Calendrier' (peuplé en place depuis T554S.TypAb)
    n_cal = populate_calendar_sheet(wb, calendar)

    # 5. Sauvegarde (crée le dossier de sortie au besoin)
    print(f"[Sauvegarde] → {output.name}")
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)

    print()
    print(f"=== Terminé : {len(data)} catégories ({n_inj} lignes injectées), "
          f"{len(pen_classes)} classes d'absences, {n_cal} codes calendrier ===")
    print(f"Sortie : {output}")
    return {"categories": len(data), "rows": n_inj,
            "classes": len(pen_classes), "calendar": n_cal,
            "output": str(output)}


def generate(client: str) -> None:
    """Génère pour un client pré-configuré dans CLIENT_CONFIGS (usage CLI)."""
    generate_from_config(CLIENT_CONFIGS[client])


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Générateur Wagetype Catalog Absences (multi-clients)")
    parser.add_argument("--client", choices=sorted(CLIENT_CONFIGS.keys()),
                        default="AKN", help="Client à traiter (défaut : AKN)")
    parser.add_argument("--all", action="store_true",
                        help="Génère le fichier pour tous les clients configurés")
    args = parser.parse_args()

    clients = sorted(CLIENT_CONFIGS.keys()) if args.all else [args.client]
    for c in clients:
        generate(c)
        if len(clients) > 1:
            print()


if __name__ == "__main__":
    main()
