#!/usr/bin/env python3
"""
Compare_PDF V4.0 - Comparateur de bulletins de paie PDF
Version avec interface graphique complète
"""
import re
import sys
import os
import threading
from collections import Counter
from datetime import datetime

# === IMPORTS GUI ===
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

# === CONFIG COULEURS ===
COLOR_ID_BG = "1F497D"
COLOR_ID_FONT = "FFFFFF"
COLOR_SUPPRIME = "FFC7CE"
COLOR_AJOUTE = "C6EFCE"
COLOR_BRUT = "FFEB9C"
COLOR_DETAIL = "FFFFCC"
COLOR_HEADER_SUB = "D9E1F2"
COLOR_PRE_BG = "1D6F42"
COLOR_PRE_SUB = "A9D08E"
COLOR_POST_BG = "C65911"
COLOR_POST_SUB = "F4B084"
COLOR_DELTA_BG = "710C04"      # Blood - Rouge sang foncé
COLOR_DELTA_SUB = "BC544B"     # Blush - Rouge pâle

# === FONCTIONS MÉTIER (inchangées) ===

def clean_val(text):
    if not text:
        return "---"
    return re.sub(r'\s+', ' ', text.strip())

def normalize_lib(text):
    """Normalise un libellé pour le matching (minuscules, espaces unifiés)"""
    if not text:
        return ""
    return re.sub(r'\s+', ' ', text.strip().lower())

def parse_values(raw_text):
    if not raw_text or raw_text == "---":
        return {"base": "---", "tx1": "---", "mt1": "---", "tx2": "---", "mt2": "---"}
    
    pattern = r'([\d.]+,\d+-?)'
    full_matches = re.findall(pattern, raw_text)
    
    result = {"base": "---", "tx1": "---", "mt1": "---", "tx2": "---", "mt2": "---"}
    
    if len(full_matches) >= 1:
        result["base"] = full_matches[0]
    if len(full_matches) >= 2:
        result["tx1"] = full_matches[1]
    if len(full_matches) >= 3:
        result["mt1"] = full_matches[2]
    if len(full_matches) >= 4:
        result["tx2"] = full_matches[3]
    if len(full_matches) >= 5:
        result["mt2"] = full_matches[4]
    
    if len(full_matches) == 2:
        result["tx1"] = "---"
        result["mt1"] = full_matches[1]
    
    return result

def parse_number(val_str):
    """Convertit une valeur string format FR en float (ex: '1.234,56-' -> -1234.56)"""
    if not val_str or val_str == "---":
        return None
    try:
        # Nettoyer la chaîne
        clean = val_str.strip()
        # Gérer le signe négatif en fin de chaîne (format SAP)
        is_negative = clean.endswith('-')
        if is_negative:
            clean = clean[:-1]
        # Supprimer les points (séparateurs de milliers)
        clean = clean.replace('.', '')
        # Remplacer la virgule par un point (séparateur décimal)
        clean = clean.replace(',', '.')
        # Convertir en float
        result = float(clean)
        return -result if is_negative else result
    except (ValueError, AttributeError):
        return None

def calculate_delta(val_pre, val_post):
    """Calcule la différence POST - PRE et retourne une string formatée"""
    num_pre = parse_number(val_pre)
    num_post = parse_number(val_post)
    
    if num_pre is None and num_post is None:
        return "---"
    if num_pre is None:
        num_pre = 0
    if num_post is None:
        num_post = 0
    
    delta = num_post - num_pre
    
    if delta == 0:
        return "0,00"
    
    # Formater en style français
    is_negative = delta < 0
    abs_delta = abs(delta)
    
    # Formater avec 2 décimales
    formatted = f"{abs_delta:,.2f}"
    # Convertir au format FR: 1,234.56 -> 1.234,56
    formatted = formatted.replace(',', ' ').replace('.', ',').replace(' ', '.')
    
    if is_negative:
        formatted = formatted + "-"
    
    return formatted

def extract_text_by_matricule(pdf_path, progress_callback=None):
    try:
        from pypdf import PdfReader
        reader = PdfReader(pdf_path)
    except ImportError:
        raise Exception("pypdf non installé. Exécutez: pip install pypdf")
    except Exception as e:
        raise Exception(f"Erreur lecture PDF: {e}")

    grouped_data = {}
    regex_strict = re.compile(
        r'(?:Matricule|Matr\.|Pers\.No\.)\s*[:.]?\s*(\d{4,10})', 
        re.IGNORECASE | re.MULTILINE
    )

    current_matricule = "INCONNU"
    total_pages = len(reader.pages)

    for i, page in enumerate(reader.pages):
        if progress_callback:
            progress_callback(i + 1, total_pages)
        
        text = page.extract_text()
        if not text:
            continue

        search_text = text.replace('\n', ' ')
        match = regex_strict.search(search_text)
        
        if match:
            new_mat = match.group(1)
            if len(new_mat) <= 10:
                current_matricule = new_mat
        
        if current_matricule not in grouped_data:
            grouped_data[current_matricule] = ""
        grouped_data[current_matricule] += "\n" + text

    return grouped_data

def parse_lines_flexible(text):
    lines = text.split('\n')
    data = {}
    key_counter = Counter()

    regex_std = re.compile(r'^\s*([A-Z0-9/]{4})\s+(.+?)\s+(-?[\d.]+,\d+-?.*)$')
    regex_no_code = re.compile(r'^\s*([A-Za-zÀ-Öà-öØ-ÿ][^:]{2,}?)\s+([-\d.,\s]+)$')

    for line in lines:
        line = line.strip()
        if not line:
            continue

        code = "????"
        libelle = ""
        values_raw = ""
        match_found = False

        m_std = regex_std.match(line)
        if m_std:
            code = m_std.group(1)
            libelle = m_std.group(2).strip()
            values_raw = m_std.group(3).strip()
            match_found = True
        
        if not match_found:
            m_no = regex_no_code.match(line)
            if m_no:
                temp_lib = m_no.group(1).strip()
                temp_val = m_no.group(2).strip()
                if any(char.isdigit() for char in temp_val) and len(temp_lib) < 60:
                    code = "????"
                    libelle = temp_lib
                    values_raw = temp_val
                    match_found = True

        if match_found:
            noise_patterns = ["Matricule", "Page", "SIRET", "URSSAF :", "IBAN", "N° SS"]
            if any(x in libelle for x in noise_patterns):
                continue

            norm_key = f"{code}|{normalize_lib(libelle)}"
            key_counter[norm_key] += 1
            occurrence = key_counter[norm_key]
            unique_key = f"{norm_key}|#{occurrence}"

            data[unique_key] = {
                "code": code,
                "lib": libelle,
                "val": values_raw,
                "occ": occurrence
            }
    
    return data

def compare_content(text_pre, text_post, show_details=False):
    d_pre = parse_lines_flexible(text_pre) if text_pre else {}
    d_post = parse_lines_flexible(text_post) if text_post else {}
    
    all_keys = sorted(list(set(d_pre.keys()) | set(d_post.keys())))
    
    causes = []
    impact_brut = []
    impact_detail = []
    
    for key in all_keys:
        i_pre = d_pre.get(key)
        i_post = d_post.get(key)
        
        code = i_pre['code'] if i_pre else i_post['code']
        lib = i_pre['lib'] if i_pre else i_post['lib']
        occ = i_pre['occ'] if i_pre else i_post['occ']
        val_pre = clean_val(i_pre['val']) if i_pre else "---"
        val_post = clean_val(i_post['val']) if i_post else "---"

        is_retro = (occ > 1)

        if i_pre and not i_post:
            if not show_details and is_retro:
                continue
            causes.append([code, lib, "SUPPRIMÉ", val_pre, val_post])
        elif not i_pre and i_post:
            if not show_details and is_retro:
                continue
            causes.append([code, lib, "AJOUTÉ", val_pre, val_post])
        elif i_pre and i_post:
            v_pre_clean = ''.join(i_pre['val'].split())
            v_post_clean = ''.join(i_post['val'].split())
            
            if v_pre_clean != v_post_clean:
                if "BRUT" in lib.upper() or "GROSS" in lib.upper():
                    impact_brut.append([code, lib, "IMPACT BRUT", val_pre, val_post])
                else:
                    if show_details:
                        impact_detail.append([code, lib, "RECALCUL", val_pre, val_post])
                    
    return causes, impact_brut, impact_detail

def generate_excel_final(global_results, filename="Comparaison-PRE-POST.xlsx"):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparaison"
    
    fill_header = PatternFill(start_color=COLOR_ID_BG, end_color=COLOR_ID_BG, fill_type="solid")
    fill_header_sub = PatternFill(start_color=COLOR_HEADER_SUB, end_color=COLOR_HEADER_SUB, fill_type="solid")
    fill_suppr = PatternFill(start_color=COLOR_SUPPRIME, end_color=COLOR_SUPPRIME, fill_type="solid")
    fill_ajout = PatternFill(start_color=COLOR_AJOUTE, end_color=COLOR_AJOUTE, fill_type="solid")
    fill_brut = PatternFill(start_color=COLOR_BRUT, end_color=COLOR_BRUT, fill_type="solid")
    fill_detail = PatternFill(start_color=COLOR_DETAIL, end_color=COLOR_DETAIL, fill_type="solid")
    fill_pre = PatternFill(start_color=COLOR_PRE_BG, end_color=COLOR_PRE_BG, fill_type="solid")
    fill_pre_sub = PatternFill(start_color=COLOR_PRE_SUB, end_color=COLOR_PRE_SUB, fill_type="solid")
    fill_post = PatternFill(start_color=COLOR_POST_BG, end_color=COLOR_POST_BG, fill_type="solid")
    fill_post_sub = PatternFill(start_color=COLOR_POST_SUB, end_color=COLOR_POST_SUB, fill_type="solid")
    fill_delta = PatternFill(start_color=COLOR_DELTA_BG, end_color=COLOR_DELTA_BG, fill_type="solid")
    fill_delta_sub = PatternFill(start_color=COLOR_DELTA_SUB, end_color=COLOR_DELTA_SUB, fill_type="solid")
    
    font_header = Font(bold=True, color=COLOR_ID_FONT, size=11)
    font_header_sub = Font(bold=True, size=10)
    font_normal = Font(name='Arial', size=10)
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers_row1 = ["Code", "Libellé", "Statut", "Valeur PRE", "", "", "", "", "Valeur POST", "", "", "", "", "DELTA", "", "", "", ""]
    for col, val in enumerate(headers_row1, 1):
        cell = ws.cell(row=1, column=col, value=val)
        if col <= 3:
            cell.fill = fill_header
        elif col <= 8:
            cell.fill = fill_pre
        elif col <= 13:
            cell.fill = fill_post
        else:
            cell.fill = fill_delta
        cell.font = font_header
        cell.alignment = align_center
        cell.border = thin_border
    
    ws.merge_cells('D1:H1')
    ws.merge_cells('I1:M1')
    ws.merge_cells('N1:R1')
    
    headers_row2 = ["", "", "", "Base", "Tx Sal", "Mt Sal", "Tx Pat", "Mt Pat", "Base", "Tx Sal", "Mt Sal", "Tx Pat", "Mt Pat", "Base", "Tx Sal", "Mt Sal", "Tx Pat", "Mt Pat"]
    for col, val in enumerate(headers_row2, 1):
        cell = ws.cell(row=2, column=col, value=val)
        if col <= 3:
            cell.fill = fill_header_sub
        elif col <= 8:
            cell.fill = fill_pre_sub
        elif col <= 13:
            cell.fill = fill_post_sub
        else:
            cell.fill = fill_delta_sub
        cell.font = font_header_sub
        cell.alignment = align_center
        cell.border = thin_border
    
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 12
    for col_letter in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']:
        ws.column_dimensions[col_letter].width = 12
    
    current_row = 3
    
    for mat_id, data in global_results.items():
        causes, brut, detail = data['causes'], data['brut'], data['detail']
        if not causes and not brut and not detail:
            continue
        
        ws.cell(row=current_row, column=1, value=f"Matricule {mat_id}")
        for col in range(1, 19):
            cell = ws.cell(row=current_row, column=col)
            if col <= 3:
                cell.fill = fill_header
            elif col <= 8:
                cell.fill = fill_pre
            elif col <= 13:
                cell.fill = fill_post
            else:
                cell.fill = fill_delta
            cell.font = font_header
            cell.border = thin_border
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=8)
        ws.merge_cells(start_row=current_row, start_column=9, end_row=current_row, end_column=13)
        ws.merge_cells(start_row=current_row, start_column=14, end_row=current_row, end_column=18)
        current_row += 1
        
        all_rows = causes + brut + detail
        
        for row_data in all_rows:
            code, lib, statut, val_pre, val_post = row_data
            
            parsed_pre = parse_values(val_pre)
            parsed_post = parse_values(val_post)
            
            # Calculer les deltas
            delta_base = calculate_delta(parsed_pre["base"], parsed_post["base"])
            delta_tx1 = calculate_delta(parsed_pre["tx1"], parsed_post["tx1"])
            delta_mt1 = calculate_delta(parsed_pre["mt1"], parsed_post["mt1"])
            delta_tx2 = calculate_delta(parsed_pre["tx2"], parsed_post["tx2"])
            delta_mt2 = calculate_delta(parsed_pre["mt2"], parsed_post["mt2"])
            
            if statut == "SUPPRIMÉ":
                fill = fill_suppr
            elif statut == "AJOUTÉ":
                fill = fill_ajout
            elif statut == "IMPACT BRUT":
                fill = fill_brut
            else:
                fill = fill_detail
            
            row_values = [
                code, lib, statut,
                parsed_pre["base"], parsed_pre["tx1"], parsed_pre["mt1"], parsed_pre["tx2"], parsed_pre["mt2"],
                parsed_post["base"], parsed_post["tx1"], parsed_post["mt1"], parsed_post["tx2"], parsed_post["mt2"],
                delta_base, delta_tx1, delta_mt1, delta_tx2, delta_mt2,
            ]
            
            for col, val in enumerate(row_values, 1):
                cell = ws.cell(row=current_row, column=col, value=val)
                cell.fill = fill
                cell.font = font_normal
                cell.border = thin_border
                if col >= 4:
                    cell.alignment = align_right
                else:
                    cell.alignment = align_left
            
            current_row += 1
        
        current_row += 1
    
    ws.freeze_panes = 'A3'
    wb.save(filename)
    return filename


# === INTERFACE GRAPHIQUE ===

class ComparePDFApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Comparateur de Bulletins de Paie PDF")
        self.root.geometry("700x500")
        self.root.minsize(700, 500)
        
        # Variables
        self.file_pre = None
        self.file_post = None
        self.mode_detail = tk.BooleanVar(value=False)
        
        # Style
        self.root.configure(bg="#f0f0f0")
        
        self.create_widgets()
        
    def create_widgets(self):
        # === FRAME PRINCIPAL ===
        main_frame = tk.Frame(self.root, bg="#f0f0f0", padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # === TITRE ===
        title_label = tk.Label(
            main_frame, 
            text="Comparateur de Bulletins de Paie", 
            font=("Arial", 16, "bold"),
            bg="#f0f0f0",
            fg="#1F497D"
        )
        title_label.pack(pady=(0, 20))
        
        # === FRAME FICHIERS (PRE et POST côte à côte) ===
        files_frame = tk.Frame(main_frame, bg="#f0f0f0")
        files_frame.pack(fill=tk.X, pady=10)
        
        # --- Colonne PRE ---
        pre_frame = tk.LabelFrame(
            files_frame, 
            text=" Fichier PRE (AVANT) ", 
            font=("Arial", 10, "bold"),
            bg="#f0f0f0",
            fg="#1D6F42",
            padx=15,
            pady=15
        )
        pre_frame.pack(side=tk.LEFT, expand=True, fill=tk.BOTH, padx=(0, 10))
        
        self.btn_pre = tk.Button(
            pre_frame, 
            text="📂 Ouvrir fichier PRE",
            font=("Arial", 11),
            bg="#1D6F42",
            fg="white",
            activebackground="#165a34",
            activeforeground="white",
            cursor="hand2",
            width=20,
            height=2,
            command=self.open_file_pre
        )
        self.btn_pre.pack(pady=(0, 10))
        
        self.label_pre = tk.Label(
            pre_frame,
            text="Aucun fichier sélectionné",
            font=("Arial", 9),
            bg="#f0f0f0",
            fg="#666666",
            wraplength=280
        )
        self.label_pre.pack()
        
        self.info_pre = tk.Label(
            pre_frame,
            text="",
            font=("Arial", 9, "italic"),
            bg="#f0f0f0",
            fg="#1D6F42"
        )
        self.info_pre.pack(pady=(5, 0))
        
        # --- Colonne POST ---
        post_frame = tk.LabelFrame(
            files_frame, 
            text=" Fichier POST (APRÈS) ", 
            font=("Arial", 10, "bold"),
            bg="#f0f0f0",
            fg="#C65911",
            padx=15,
            pady=15
        )
        post_frame.pack(side=tk.RIGHT, expand=True, fill=tk.BOTH, padx=(10, 0))
        
        self.btn_post = tk.Button(
            post_frame, 
            text="📂 Ouvrir fichier POST",
            font=("Arial", 11),
            bg="#C65911",
            fg="white",
            activebackground="#a34a0e",
            activeforeground="white",
            cursor="hand2",
            width=20,
            height=2,
            command=self.open_file_post
        )
        self.btn_post.pack(pady=(0, 10))
        
        self.label_post = tk.Label(
            post_frame,
            text="Aucun fichier sélectionné",
            font=("Arial", 9),
            bg="#f0f0f0",
            fg="#666666",
            wraplength=280
        )
        self.label_post.pack()
        
        self.info_post = tk.Label(
            post_frame,
            text="",
            font=("Arial", 9, "italic"),
            bg="#f0f0f0",
            fg="#C65911"
        )
        self.info_post.pack(pady=(5, 0))
        
        # === OPTIONS ===
        options_frame = tk.Frame(main_frame, bg="#f0f0f0")
        options_frame.pack(fill=tk.X, pady=20)
        
        self.check_detail = tk.Checkbutton(
            options_frame,
            text="Mode DÉTAILLÉ (afficher toutes les différences)",
            variable=self.mode_detail,
            font=("Arial", 10),
            bg="#f0f0f0",
            activebackground="#f0f0f0"
        )
        self.check_detail.pack()
        
        # === BARRE DE PROGRESSION ===
        self.progress_frame = tk.Frame(main_frame, bg="#f0f0f0")
        self.progress_frame.pack(fill=tk.X, pady=10)
        
        self.progress_label = tk.Label(
            self.progress_frame,
            text="",
            font=("Arial", 9),
            bg="#f0f0f0",
            fg="#666666"
        )
        self.progress_label.pack()
        
        self.progress_bar = ttk.Progressbar(
            self.progress_frame,
            mode='determinate',
            length=400
        )
        self.progress_bar.pack(pady=(5, 0))
        self.progress_bar.pack_forget()  # Caché par défaut
        
        # === BOUTON LANCER ===
        self.btn_compare = tk.Button(
            main_frame,
            text="🚀 Lancer la comparaison",
            font=("Arial", 12, "bold"),
            bg="#1F497D",
            fg="white",
            activebackground="#163a5c",
            activeforeground="white",
            disabledforeground="#999999",
            cursor="hand2",
            width=25,
            height=2,
            state=tk.DISABLED,
            command=self.run_comparison
        )
        self.btn_compare.pack(pady=10)
        
        # === STATUS ===
        self.status_label = tk.Label(
            main_frame,
            text="Sélectionnez les deux fichiers PDF pour commencer",
            font=("Arial", 9, "italic"),
            bg="#f0f0f0",
            fg="#666666"
        )
        self.status_label.pack(pady=(10, 0))

    def open_file_pre(self):
        filepath = filedialog.askopenfilename(
            title="Sélectionner le fichier PRE (AVANT)",
            filetypes=[("Fichiers PDF", "*.pdf")]
        )
        if filepath:
            self.file_pre = filepath
            filename = os.path.basename(filepath)
            self.label_pre.config(text=filename, fg="#1D6F42")
            
            # Compter les pages
            try:
                from pypdf import PdfReader
                reader = PdfReader(filepath)
                nb_pages = len(reader.pages)
                self.info_pre.config(text=f"({nb_pages} pages)")
            except Exception:
                self.info_pre.config(text="")
            
            self.check_ready()
    
    def open_file_post(self):
        filepath = filedialog.askopenfilename(
            title="Sélectionner le fichier POST (APRÈS)",
            filetypes=[("Fichiers PDF", "*.pdf")]
        )
        if filepath:
            self.file_post = filepath
            filename = os.path.basename(filepath)
            self.label_post.config(text=filename, fg="#C65911")
            
            # Compter les pages
            try:
                from pypdf import PdfReader
                reader = PdfReader(filepath)
                nb_pages = len(reader.pages)
                self.info_post.config(text=f"({nb_pages} pages)")
            except Exception:
                self.info_post.config(text="")
            
            self.check_ready()
    
    def check_ready(self):
        """Active le bouton si les deux fichiers sont sélectionnés"""
        if self.file_pre and self.file_post:
            if os.path.normpath(self.file_pre) == os.path.normpath(self.file_post):
                self.btn_compare.config(state=tk.DISABLED, bg="#cccccc")
                self.status_label.config(
                    text="⚠ Les fichiers PRE et POST sont identiques", fg="#C65911"
                )
            else:
                self.btn_compare.config(state=tk.NORMAL, bg="#1F497D")
                self.status_label.config(
                    text="✓ Prêt ! Cliquez sur 'Lancer la comparaison'", fg="#1D6F42"
                )
        else:
            self.btn_compare.config(state=tk.DISABLED, bg="#cccccc")

    def update_progress(self, current, total, phase=""):
        """Met à jour la barre de progression (thread-safe via root.after)"""
        self.root.after(0, self._do_update_progress, current, total, phase)

    def _do_update_progress(self, current, total, phase):
        percent = (current / total) * 100
        self.progress_bar['value'] = percent
        self.progress_label.config(text=f"{phase} : {current}/{total} pages")

    def run_comparison(self):
        """Lance la comparaison dans un thread séparé"""
        # Désactiver les boutons pendant le traitement
        self.btn_pre.config(state=tk.DISABLED)
        self.btn_post.config(state=tk.DISABLED)
        self.btn_compare.config(state=tk.DISABLED, text="⏳ Traitement en cours...")

        # Lire le mode détaillé depuis le thread principal (Tkinter var)
        self._show_details = self.mode_detail.get()

        # Afficher la barre de progression
        self.progress_bar.pack(pady=(5, 0))
        self.progress_bar['value'] = 0

        # Lancer dans un thread
        thread = threading.Thread(target=self.do_comparison, daemon=True)
        thread.start()

    def do_comparison(self):
        """Effectue l'extraction et la comparaison (dans un thread secondaire)"""
        try:
            # Extraction PRE
            self.root.after(0, lambda: self.progress_label.config(
                text="Lecture du fichier PRE..."
            ))

            data_pre = extract_text_by_matricule(
                self.file_pre,
                lambda c, t: self.update_progress(c, t, "PRE")
            )

            # Extraction POST
            self.root.after(0, lambda: self.progress_label.config(
                text="Lecture du fichier POST..."
            ))

            data_post = extract_text_by_matricule(
                self.file_post,
                lambda c, t: self.update_progress(c, t, "POST")
            )

            # Collecter les avertissements
            warnings = []

            # Détection PDF scanné (aucun texte extractible)
            text_pre_total = sum(len(v) for v in data_pre.values())
            text_post_total = sum(len(v) for v in data_post.values())
            if text_pre_total == 0:
                warnings.append(
                    "Le fichier PRE semble être un PDF scanné "
                    "(aucun texte extractible)."
                )
            if text_post_total == 0:
                warnings.append(
                    "Le fichier POST semble être un PDF scanné "
                    "(aucun texte extractible)."
                )

            # Avertissement matricule INCONNU
            if "INCONNU" in data_pre:
                warnings.append(
                    "PRE : certaines pages n'ont pas de matricule identifiable."
                )
            if "INCONNU" in data_post:
                warnings.append(
                    "POST : certaines pages n'ont pas de matricule identifiable."
                )

            # Comparaison
            self.root.after(0, lambda: self.progress_label.config(
                text="Comparaison en cours..."
            ))
            self.root.after(0, lambda: self.progress_bar.configure(value=50))

            global_results = {}
            all_mats = sorted(list(set(data_pre.keys()) | set(data_post.keys())))

            for mat in all_mats:
                if mat == "INCONNU":
                    continue
                causes, brut, detail = compare_content(
                    data_pre.get(mat, ""),
                    data_post.get(mat, ""),
                    show_details=self._show_details
                )
                if causes or brut or detail:
                    global_results[mat] = {
                        "causes": causes, "brut": brut, "detail": detail
                    }

            # Passer au thread principal pour les dialogues GUI
            self.root.after(0, self._finalize_comparison, global_results, warnings)

        except Exception as e:
            self.root.after(0, self._handle_error, str(e))

    def _finalize_comparison(self, global_results, warnings):
        """Finalise la comparaison sur le thread principal (dialogues GUI)"""
        try:
            self.progress_label.config(text="Génération du fichier Excel...")
            self.progress_bar['value'] = 80

            # Afficher les avertissements
            if warnings:
                messagebox.showwarning(
                    "Avertissements",
                    "\n\n".join(warnings)
                )

            if global_results:
                # Nom de fichier avec timestamp
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

                output_file = filedialog.asksaveasfilename(
                    title="Enregistrer le résultat",
                    defaultextension=".xlsx",
                    filetypes=[("Fichier Excel", "*.xlsx")],
                    initialfile=f"Comparaison-PRE-POST_{timestamp}.xlsx"
                )

                if output_file:
                    generate_excel_final(global_results, output_file)

                    self.progress_bar['value'] = 100

                    nb_ecarts = len(global_results)
                    result = messagebox.askyesno(
                        "Terminé !",
                        f"✓ {nb_ecarts} matricule(s) avec écarts détectés.\n\n"
                        f"Fichier créé :\n{output_file}\n\n"
                        "Voulez-vous ouvrir le fichier Excel ?"
                    )

                    if result:
                        try:
                            os.startfile(output_file)
                        except AttributeError:
                            import subprocess
                            subprocess.run(['xdg-open', output_file], check=False)
            else:
                messagebox.showinfo(
                    "Résultat",
                    "✓ R.A.S - Aucune différence trouvée entre les deux fichiers."
                )
        finally:
            self._reset_ui()

    def _handle_error(self, error_msg):
        """Gère les erreurs sur le thread principal"""
        messagebox.showerror(
            "Erreur", f"Une erreur s'est produite :\n{error_msg}"
        )
        self._reset_ui()

    def _reset_ui(self):
        """Réinitialise l'interface après comparaison"""
        self.btn_pre.config(state=tk.NORMAL)
        self.btn_post.config(state=tk.NORMAL)
        self.btn_compare.config(state=tk.NORMAL, text="🚀 Lancer la comparaison")
        self.progress_bar.pack_forget()
        self.progress_label.config(text="")
        self.status_label.config(
            text="✓ Prêt pour une nouvelle comparaison", fg="#1D6F42"
        )


# === POINT D'ENTRÉE ===

def main():
    # Vérifier les dépendances
    try:
        from pypdf import PdfReader
    except ImportError:
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror(
            "Dépendance manquante",
            "Le module 'pypdf' n'est pas installé.\n\n"
            "Exécutez : pip install pypdf"
        )
        sys.exit(1)
    
    try:
        import openpyxl
    except ImportError:
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror(
            "Dépendance manquante",
            "Le module 'openpyxl' n'est pas installé.\n\n"
            "Exécutez : pip install openpyxl"
        )
        sys.exit(1)
    
    # Lancer l'application
    root = tk.Tk()
    app = ComparePDFApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
