"""
Scanner de tickets SAP pour problèmes DSN - Source 2 : par client
v1.0 - Scan les ~1322 fichiers Excel dans le répertoire SAP Ticket
"""
APP_VERSION = "1.0"

import openpyxl
import os
import json
import re
import sys
from pathlib import Path
from collections import defaultdict
from datetime import datetime

TICKETS_ROOT = r"E:\Dossier Manuel (CV, taf, dev etc)\NGA\Travail (NGA et autres)\SAP Ticket\SAP Ticket"
OUTPUT_DIR = r"D:\Program\ClaudeCode\SAP_DSN"

DSN_KEYWORDS_STRONG = [
    'dsn', 'bloc 78', 'bloc 23', 'bloc 50', 'bloc 53', 'bloc 54', 'bloc 81',
    'bloc 7 ', 'bloc 20', 'bloc 22', 'bloc 30', 'bloc 40', 'bloc 44',
    'bloc 51', 'bloc 56', 'bloc 60', 'bloc 65', 'bloc 70', 'bloc 79',
    'bloc 82', 'bloc 83', 'bloc 84', 'bloc 85', 'bloc 86',
    's21.g00', 's20.g00', 's10.g00', 's89.g00',
    'rpldsnf0', 'dsncheck', 'dsn check',
    'déclaration sociale nominative',
    'taux pas', 'taux at', 'prélèvement à la source',
    'urssaf', 'agirc', 'arrco', 'prevoyance',
    'cotisation', 'net social', 'net imposable',
    'signalement', 'annule et remplace', 'annule-remplace',
    'fractionn', 'néant', 'norme dsn', 'cahier technique',
]

DSN_FILENAME_KEYWORDS = [
    'dsn', 'bloc 78', 'bloc 23', 'bloc 50', 'bloc 53', 'bloc 54', 'bloc 81',
    'bloc 7 ', 'bloc7', 'bloc 20', 'bloc 22', 'bloc 30', 'bloc 44',
    'bloc 51', 'bloc 56', 'bloc 60', 'bloc 65', 'bloc 70', 'bloc 79',
    'bloc 82', 'bloc 83', 'bloc 84', 'bloc 85', 'bloc 86',
    'taux pas', 'taux at', 'prelevement', 'prevoyance',
    'cotisation', 'urssaf', 'net social', 'quotite',
    'rpldsnf', 'handicap',
]

DSN_SAP_TABLES = [
    'T5F99', 'T5F1P', 'T5FGRP', 'T5FADR', 'T5FDSNPAS', 'T596',
    'T5F3C', 'T5F3D', 'T5F3E', 'T5F3F', 'T5F3G', 'T5F3H',
    'T5F99F0', 'T5FBL', 'T536A', 'T5FDL', 'T5FDS', 'V_T5F',
    'T511K', 'T512W', 'T512T', 'V_512W', 'T5F4',
    'IT0001', 'IT0002', 'IT0006', 'IT0007', 'IT0008', 'IT0009',
    'IT0014', 'IT0015', 'IT0041', 'IT0105', 'IT0185', 'IT0207',
    'IT0208', 'IT0209', 'IT0210', 'IT0211', 'IT3331', 'IT3332',
    'PA0001', 'PA0002', 'PA0006', 'PA0007', 'PA0008', 'PA0009',
    'PA0014', 'PA0015', 'PA0041', 'PA0105', 'PA0185', 'PA0207',
    'PA0208', 'PA0209', 'PA0210', 'PA0211', 'PA3331', 'PA3332',
    'HRPY_RGDIR', 'PCL2', 'RT', 'CRT', 'CT',
]

# Mapping chemin -> client
CLIENTS = ['ABBVIE', 'AKZO NOBEL', 'ALCON', 'ASTELLAS', 'BUNGE',
           'CORNING', 'LEO PHARMA', 'LONZA', 'RECKITT']


def extract_client(filepath):
    for c in CLIENTS:
        if c in filepath.upper():
            return c
    return 'UNKNOWN'


def extract_text_from_sheet(ws, max_rows=200):
    texts = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        for cell in row:
            if cell is not None:
                val = str(cell).strip()
                if val and len(val) > 1:
                    texts.append(val)
    return texts


def is_dsn_related(filename, all_texts):
    fname_lower = filename.lower()
    for kw in DSN_FILENAME_KEYWORDS:
        if kw in fname_lower:
            return True, f"filename:{kw}"
    combined = ' '.join(all_texts).lower()
    for kw in DSN_KEYWORDS_STRONG:
        if kw in combined:
            return True, f"content:{kw}"
    return False, ""


def extract_ticket_info(filepath):
    result = {
        'file': str(filepath),
        'filename': os.path.basename(filepath),
        'client': extract_client(str(filepath)),
        'sheets': [],
        'problem': '',
        'solution': '',
        'tables_mentioned': [],
        'blocs_mentioned': [],
        'dsn_codes': [],
    }
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        result['sheets'] = wb.sheetnames
        all_texts = []
        sheet_contents = {}
        for sname in wb.sheetnames:
            try:
                ws = wb[sname]
                texts = extract_text_from_sheet(ws)
                all_texts.extend(texts)
                sheet_contents[sname] = texts
            except Exception:
                pass

        all_text = ' '.join(all_texts)

        problem_sheets = ['INC FORM', 'DFCT FORM', 'Sheet1', 'Analyse',
                         'HRO Screeshots Request', 'Test Evidence Form',
                         'Test Evidence', 'Description']
        problem_texts = []
        for ps in problem_sheets:
            if ps in sheet_contents:
                problem_texts.extend(sheet_contents[ps])
        result['problem'] = ' | '.join(problem_texts[:50])

        solution_sheets = ['Change Logs', 'AMO', 'AMO Screeshots QAS',
                          'AMO Screenshots', 'Analyse', 'Resolution']
        solution_texts = []
        for ss in solution_sheets:
            if ss in sheet_contents:
                solution_texts.extend(sheet_contents[ss])
        result['solution'] = ' | '.join(solution_texts[:50])

        combined = all_text.upper()
        for table in DSN_SAP_TABLES:
            if table.upper() in combined:
                result['tables_mentioned'].append(table)

        bloc_pattern = re.findall(r'bloc\s*(\d+)', combined.lower())
        result['blocs_mentioned'] = list(set(bloc_pattern))

        dsn_codes = re.findall(r'S\d{2}\.G\d{2}\.\d{2}(?:\.\d{3})?', combined)
        result['dsn_codes'] = list(set(dsn_codes))

        # Pour la détection DSN
        result['_all_text'] = all_text

        wb.close()
    except Exception as e:
        result['error'] = str(e)

    return result


def scan_all_tickets():
    all_files = []
    for root, dirs, files in os.walk(TICKETS_ROOT):
        for f in files:
            if f.endswith(('.xlsx', '.xlsm', '.xls')) and not f.startswith('~$'):
                all_files.append(os.path.join(root, f))

    print(f"Fichiers Excel trouvés: {len(all_files)}")

    dsn_tickets = []
    non_dsn_count = 0
    error_count = 0

    for idx, fpath in enumerate(all_files):
        if (idx + 1) % 50 == 0:
            print(f"  Progression: {idx+1}/{len(all_files)} ({len(dsn_tickets)} DSN trouvés)")

        try:
            info = extract_ticket_info(fpath)

            if 'error' in info:
                error_count += 1
                is_dsn, match = is_dsn_related(info['filename'], [])
                if is_dsn:
                    info['dsn_match'] = match
                    info['dsn_from_filename_only'] = True
                    dsn_tickets.append(info)
                continue

            is_dsn, match = is_dsn_related(info['filename'], [info.get('_all_text', '')])
            if is_dsn:
                info['dsn_match'] = match
                info.pop('_all_text', None)
                dsn_tickets.append(info)
            else:
                non_dsn_count += 1
                info.pop('_all_text', None)

        except Exception as e:
            error_count += 1

    print(f"\nRésultat final:")
    print(f"  Total fichiers scannés: {len(all_files)}")
    print(f"  Tickets DSN trouvés: {len(dsn_tickets)}")
    print(f"  Tickets non-DSN: {non_dsn_count}")
    print(f"  Erreurs: {error_count}")

    return dsn_tickets, len(all_files), error_count


def main():
    print(f"=== Scanner DSN Tickets (par client) v{APP_VERSION} ===")
    print(f"Début: {datetime.now().strftime('%H:%M:%S')}")
    print(f"Source: {TICKETS_ROOT}")
    print()

    dsn_tickets, total_files, error_count = scan_all_tickets()

    # Sauvegarder
    raw_output = os.path.join(OUTPUT_DIR, 'dsn_tickets_clients_raw.json')
    for t in dsn_tickets:
        t.pop('_all_text', None)
    with open(raw_output, 'w', encoding='utf-8') as f:
        json.dump(dsn_tickets, f, ensure_ascii=False, indent=2, default=str)
    print(f"\nTickets DSN sauvegardés: {raw_output}")

    # Stats par client
    client_stats = defaultdict(int)
    for t in dsn_tickets:
        client_stats[t.get('client', 'UNKNOWN')] += 1
    print("\nTickets DSN par client:")
    for client, count in sorted(client_stats.items(), key=lambda x: x[1], reverse=True):
        print(f"  {client}: {count}")

    print(f"\nFin: {datetime.now().strftime('%H:%M:%S')}")


if __name__ == '__main__':
    main()
