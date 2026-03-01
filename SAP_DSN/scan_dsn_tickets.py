"""
Scanner de tickets SAP pour problèmes DSN
v1.0 - Scan les ~1888 fichiers Excel dans le répertoire des tickets Julio
Identifie les tickets liés à la DSN et extrait les problèmes + solutions
"""
APP_VERSION = "1.0"

import openpyxl
import os
import json
import re
import sys
import traceback
from pathlib import Path
from collections import defaultdict
from datetime import datetime

# Répertoire source des tickets
TICKETS_ROOT = r"E:\Dossier Manuel (CV, taf, dev etc)\NGA\Travail (NGA et autres)\Tous les TE de Julio"
OUTPUT_DIR = r"D:\Program\ClaudeCode\SAP_DSN"

# Mots-clés DSN pour détecter les tickets pertinents
DSN_KEYWORDS_STRONG = [
    'dsn', 'bloc 78', 'bloc 23', 'bloc 50', 'bloc 53', 'bloc 54', 'bloc 81',
    'bloc 7 ', 'bloc 20', 'bloc 22', 'bloc 30', 'bloc 40', 'bloc 44',
    'bloc 51', 'bloc 56', 'bloc 60', 'bloc 65', 'bloc 70', 'bloc 79',
    'bloc 82', 'bloc 83', 'bloc 84', 'bloc 85', 'bloc 86',
    's21.g00', 's20.g00', 's10.g00', 's89.g00',
    'rpldsnf0', 'dsncheck', 'dsn check',
    'déclaration sociale nominative',
    'taux pas', 'taux at', 'prélèvement à la source',
    'urssaf', 'agirc', 'arrco', 'prevoyance',
    'cotisation', 'net social', 'net imposable',
    'signalement', 'annule et remplace', 'annule-remplace',
    'fractionn', 'néant', 'norme dsn', 'cahier technique',
]

# Mots-clés qui dans le nom de fichier suffisent à identifier un ticket DSN
DSN_FILENAME_KEYWORDS = [
    'dsn', 'bloc 78', 'bloc 23', 'bloc 50', 'bloc 53', 'bloc 54', 'bloc 81',
    'bloc 7 ', 'bloc7', 'bloc 20', 'bloc 22', 'bloc 30', 'bloc 44',
    'bloc 51', 'bloc 56', 'bloc 60', 'bloc 65', 'bloc 70', 'bloc 79',
    'bloc 82', 'bloc 83', 'bloc 84', 'bloc 85', 'bloc 86',
    'taux pas', 'taux at', 'prelevement', 'prevoyance',
    'cotisation', 'urssaf', 'net social', 'quotite',
    'rpldsnf', 'handicap',
]

# Tables SAP connues liées à la DSN (EuHReka)
DSN_SAP_TABLES = [
    'T5F99', 'T5F1P', 'T5FGRP', 'T5FADR', 'T5FDSNPAS', 'T596',
    'T5F3C', 'T5F3D', 'T5F3E', 'T5F3F', 'T5F3G', 'T5F3H',
    'T5F99F0', 'T5FBL', 'T536A', 'T5FDL', 'T5FDS', 'V_T5F',
    'T511K', 'T512W', 'T512T', 'V_512W', 'T5F4',
    'IT0001', 'IT0002', 'IT0006', 'IT0007', 'IT0008', 'IT0009',
    'IT0014', 'IT0015', 'IT0041', 'IT0105', 'IT0185', 'IT0207',
    'IT0208', 'IT0209', 'IT0210', 'IT0211', 'IT3331', 'IT3332',
    'PA0001', 'PA0002', 'PA0006', 'PA0007', 'PA0008', 'PA0009',
    'PA0014', 'PA0015', 'PA0041', 'PA0105', 'PA0185', 'PA0207',
    'PA0208', 'PA0209', 'PA0210', 'PA0211', 'PA3331', 'PA3332',
    'HRPY_RGDIR', 'PCL2', 'RT', 'CRT', 'CT',
]


def extract_text_from_sheet(ws, max_rows=200):
    """Extrait tout le texte non-vide d'un onglet Excel."""
    texts = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        for cell in row:
            if cell is not None:
                val = str(cell).strip()
                if val and len(val) > 1:
                    texts.append(val)
    return texts


def is_dsn_related(filename, all_texts):
    """Détermine si le ticket est lié à la DSN."""
    fname_lower = filename.lower()

    # Vérifier le nom de fichier
    for kw in DSN_FILENAME_KEYWORDS:
        if kw in fname_lower:
            return True, f"filename:{kw}"

    # Vérifier le contenu
    combined = ' '.join(all_texts).lower()
    for kw in DSN_KEYWORDS_STRONG:
        if kw in combined:
            return True, f"content:{kw}"

    return False, ""


def extract_ticket_info(filepath):
    """Extrait les informations d'un fichier Excel de ticket."""
    result = {
        'file': str(filepath),
        'filename': os.path.basename(filepath),
        'sheets': [],
        'problem': '',
        'solution': '',
        'tables_mentioned': [],
        'blocs_mentioned': [],
        'dsn_codes': [],
        'all_text': '',
    }

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        result['sheets'] = wb.sheetnames

        all_texts = []
        sheet_contents = {}

        for sname in wb.sheetnames:
            try:
                ws = wb[sname]
                texts = extract_text_from_sheet(ws)
                all_texts.extend(texts)
                sheet_contents[sname] = texts
            except Exception:
                pass

        result['all_text'] = ' '.join(all_texts)

        # Extraire la description du problème
        # Chercher dans les onglets les plus probables
        problem_sheets = ['INC FORM', 'DFCT FORM', 'Sheet1', 'Analyse',
                         'HRO Screeshots Request', 'Test Evidence Form',
                         'Test Evidence', 'Description']
        problem_texts = []
        for ps in problem_sheets:
            if ps in sheet_contents:
                problem_texts.extend(sheet_contents[ps])
        result['problem'] = ' | '.join(problem_texts[:50])

        # Extraire la solution depuis Change Logs ou AMO
        solution_sheets = ['Change Logs', 'AMO', 'AMO Screeshots QAS',
                          'AMO Screenshots', 'Analyse', 'Resolution']
        solution_texts = []
        for ss in solution_sheets:
            if ss in sheet_contents:
                solution_texts.extend(sheet_contents[ss])
        result['solution'] = ' | '.join(solution_texts[:50])

        # Détecter les tables SAP mentionnées
        combined = result['all_text'].upper()
        for table in DSN_SAP_TABLES:
            if table.upper() in combined:
                result['tables_mentioned'].append(table)

        # Détecter les blocs DSN mentionnés
        bloc_pattern = re.findall(r'bloc\s*(\d+)', combined.lower())
        result['blocs_mentioned'] = list(set(bloc_pattern))

        # Détecter les codes DSN (S21.G00.xx.xxx etc)
        dsn_codes = re.findall(r'S\d{2}\.G\d{2}\.\d{2}(?:\.\d{3})?', combined)
        result['dsn_codes'] = list(set(dsn_codes))

        wb.close()

    except Exception as e:
        result['error'] = str(e)

    return result


def scan_all_tickets():
    """Scan tous les fichiers Excel dans le répertoire des tickets."""
    all_files = []
    for root, dirs, files in os.walk(TICKETS_ROOT):
        for f in files:
            if f.endswith(('.xlsx', '.xlsm', '.xls')) and not f.startswith('~$'):
                all_files.append(os.path.join(root, f))

    print(f"Fichiers Excel trouvés: {len(all_files)}")

    dsn_tickets = []
    non_dsn_count = 0
    error_count = 0

    for idx, fpath in enumerate(all_files):
        if (idx + 1) % 50 == 0:
            print(f"  Progression: {idx+1}/{len(all_files)} ({len(dsn_tickets)} DSN trouvés)")

        try:
            info = extract_ticket_info(fpath)

            if 'error' in info:
                error_count += 1
                # Quand même vérifier le nom du fichier
                is_dsn, match = is_dsn_related(info['filename'], [])
                if is_dsn:
                    info['dsn_match'] = match
                    info['dsn_from_filename_only'] = True
                    dsn_tickets.append(info)
                continue

            is_dsn, match = is_dsn_related(info['filename'], [info['all_text']])
            if is_dsn:
                info['dsn_match'] = match
                # Retirer all_text pour alléger le JSON (on garde problem/solution)
                info['all_text'] = info['all_text'][:2000]
                dsn_tickets.append(info)
            else:
                non_dsn_count += 1

        except Exception as e:
            error_count += 1

    print(f"\nRésultat final:")
    print(f"  Total fichiers scannés: {len(all_files)}")
    print(f"  Tickets DSN trouvés: {len(dsn_tickets)}")
    print(f"  Tickets non-DSN: {non_dsn_count}")
    print(f"  Erreurs: {error_count}")

    return dsn_tickets


def analyze_dsn_tickets(tickets):
    """Analyse les tickets DSN pour trouver patterns et solutions."""

    # Compteur de blocs
    bloc_counter = defaultdict(int)
    # Compteur de tables
    table_counter = defaultdict(int)
    # Compteur de codes DSN
    code_counter = defaultdict(int)
    # Problèmes par bloc
    problems_by_bloc = defaultdict(list)
    # Solutions par bloc
    solutions_by_bloc = defaultdict(list)
    # Tickets par année
    tickets_by_year = defaultdict(int)

    for t in tickets:
        # Année
        for year in ['2019', '2020', '2021', '2022', '2023']:
            if year in t['file']:
                tickets_by_year[year] += 1
                break

        # Blocs
        for bloc in t.get('blocs_mentioned', []):
            bloc_counter[bloc] += 1
            if t.get('problem'):
                problems_by_bloc[bloc].append({
                    'file': t['filename'],
                    'problem': t['problem'][:500],
                    'solution': t.get('solution', '')[:500],
                })

        # Tables
        for table in t.get('tables_mentioned', []):
            table_counter[table] += 1

        # Codes DSN
        for code in t.get('dsn_codes', []):
            code_counter[code] += 1

    analysis = {
        'total_dsn_tickets': len(tickets),
        'tickets_by_year': dict(tickets_by_year),
        'top_blocs': dict(sorted(bloc_counter.items(), key=lambda x: x[1], reverse=True)[:20]),
        'top_tables': dict(sorted(table_counter.items(), key=lambda x: x[1], reverse=True)[:30]),
        'top_dsn_codes': dict(sorted(code_counter.items(), key=lambda x: x[1], reverse=True)[:30]),
        'problems_by_bloc': {k: v[:5] for k, v in sorted(problems_by_bloc.items(),
                            key=lambda x: len(x[1]), reverse=True)[:15]},
    }

    return analysis


def main():
    print(f"=== Scanner DSN Tickets v{APP_VERSION} ===")
    print(f"Début: {datetime.now().strftime('%H:%M:%S')}")
    print(f"Source: {TICKETS_ROOT}")
    print()

    # Phase 1: Scan
    print("Phase 1: Scan de tous les fichiers Excel...")
    dsn_tickets = scan_all_tickets()

    # Sauvegarder les tickets DSN bruts
    raw_output = os.path.join(OUTPUT_DIR, 'dsn_tickets_raw.json')
    # Nettoyer pour JSON serialization
    for t in dsn_tickets:
        t.pop('all_text', None)
    with open(raw_output, 'w', encoding='utf-8') as f:
        json.dump(dsn_tickets, f, ensure_ascii=False, indent=2, default=str)
    print(f"Tickets DSN sauvegardés: {raw_output}")

    # Phase 2: Analyse
    print("\nPhase 2: Analyse des tickets DSN...")
    analysis = analyze_dsn_tickets(dsn_tickets)

    analysis_output = os.path.join(OUTPUT_DIR, 'dsn_analysis.json')
    with open(analysis_output, 'w', encoding='utf-8') as f:
        json.dump(analysis, f, ensure_ascii=False, indent=2, default=str)
    print(f"Analyse sauvegardée: {analysis_output}")

    # Phase 3: Résumé textuel
    print("\n" + "="*80)
    print("RÉSUMÉ DE L'ANALYSE DSN")
    print("="*80)

    print(f"\nTickets DSN trouvés: {analysis['total_dsn_tickets']}")
    print(f"\nPar année: {analysis['tickets_by_year']}")

    print(f"\nTop blocs DSN problématiques:")
    for bloc, count in list(analysis['top_blocs'].items())[:15]:
        print(f"  Bloc {bloc}: {count} tickets")

    print(f"\nTop tables SAP utilisées:")
    for table, count in list(analysis['top_tables'].items())[:20]:
        print(f"  {table}: {count} tickets")

    print(f"\nTop codes DSN:")
    for code, count in list(analysis['top_dsn_codes'].items())[:15]:
        print(f"  {code}: {count} tickets")

    print(f"\nFin: {datetime.now().strftime('%H:%M:%S')}")


if __name__ == '__main__':
    main()
